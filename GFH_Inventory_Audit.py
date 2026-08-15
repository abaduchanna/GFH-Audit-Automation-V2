from __future__ import annotations

import sys
from datetime import date
from header_manager import FixedHeaderManager
from logo_handler import LogoHandler

# Developed by Abad Umair Channa | Copyright © {date.today().year} | All rights reserved.
"""
GFH Telecom LLC Inventory Audit v27

What this app does
1. Upload Inventory_Count_Result_Details.xlsx.
2. Upload Employee_Time_Sheet.xlsx.
3. Rebuilds the GFH Inventory Status / Inventory Count Results workflow:
   - Reads serial-level inventory count results.
   - Filters Inventory_Count_Result_Details to the latest Created Date/Time per Store + Created By.
   - Deduplicates by (Store, IMEI) keeping the latest record.
   - Treats non-matched serialized rows as variances.
   - Pulls District and Rep Name from Employee_Time_Sheet by Store.
   - Merges Arizona - D1, Arizona - D2, and Arizona into Arizona.
   - Skips SIM, eSIM, SIM Card, SIM Kit, and similar SIM products from UI and sending.
4. Shows variances in a GUI with District, Store, Product, IMEI, Status, Rep Name, and Checkbox.
5. Checkbox marks the variance as cleared/resolved. Cleared rows are skipped in pending auto-send.
6. Sends selected or pending variances as PNG images to WhatsApp Desktop.
7. Send mode options:
   - District: one image batch per district.
   - Store: one image batch per store, routed to that store's district WhatsApp group.
   - Sales Rep: one image batch per sales rep per district, routed to that district WhatsApp group.
8. Opens with a blank UI and loads data only after both file paths are provided and Load Variances is clicked.
9. Keeps a SQLite log of cleared, pending, and sent variance status.

Recommended install on Windows:
    py -m pip install pillow pyautogui pyperclip pywin32 pygetwindow openpyxl

Store list and employee phone numbers are stored in the local SQLite database. Store list and employees can be imported from XLSX. WhatsApp group names are saved in the database and configurable per district. WhatsApp sends captions together with images. Final district result sending is available per district. Starting message and 3-reminder sending for uncleared variances is supported.

Run:
    Double click the .pyw file to run without a command window.

WhatsApp Desktop notes:
- Keep WhatsApp Desktop installed and logged in before sending.
- Default search shortcut is Ctrl+F. If your WhatsApp uses Ctrl+K for chat search,
  change SEARCH_SHORTCUT below to "ctrl+k".
- District group routing uses the WhatsApp group names saved in the District DMs tab.
"""

# ── Auto-install any missing pip packages ───────────────────────────────────
def _auto_install_packages() -> None:
    """Try every available installer until all packages are present."""
    REQUIRED = [
        ("PIL",         "pillow"),
        ("openpyxl",    "openpyxl"),
        ("pyautogui",   "pyautogui"),
        ("pyperclip",   "pyperclip"),
        ("win32api",    "pywin32"),
        ("pygetwindow", "pygetwindow"),
    ]
    missing_pip = []
    for import_name, pip_name in REQUIRED:
        try:
            __import__(import_name)
        except ImportError:
            missing_pip.append(pip_name)
    if not missing_pip:
        return

    import subprocess as _sp
    import sys as _sys
    pkgs = missing_pip

    # Show a quick tk splash so the user knows something is happening
    try:
        import tkinter as _tk
        from theme_manager import ThemeManager, apply_theme_to_window, get_copyright_year
        _splash = _tk.Tk()
        _splash.title("GFH Inventory Audit — Installing packages…")
        _splash.geometry("540x90")
        _splash.resizable(False, False)
        _lbl = _tk.Label(
            _splash,
            text=f"Installing: {', '.join(pkgs)}\nPlease wait — this runs once only…",
            font=("Segoe UI", 11), pady=18,
        )
        _lbl.pack()
        _splash.update()
    except Exception:
        _splash = None

    installers = [
        [_sys.executable, "-m", "pip", "install", "--quiet"] + pkgs,
        ["py", "-m", "pip", "install", "--quiet"] + pkgs,
        ["python", "-m", "pip", "install", "--quiet"] + pkgs,
        ["python3", "-m", "pip", "install", "--quiet"] + pkgs,
        ["pip", "install", "--quiet"] + pkgs,
        ["pip3", "install", "--quiet"] + pkgs,
    ]
    success = False
    for cmd in installers:
        try:
            r = _sp.run(cmd, capture_output=True, timeout=180)
            if r.returncode == 0:
                success = True
                break
        except Exception:
            continue

    if _splash:
        _splash.destroy()

    if not success:
        try:
            import tkinter as _tk2
            from theme_manager import ThemeManager, apply_theme_to_window, get_copyright_year
            import tkinter.messagebox as _mb
            _r2 = _tk2.Tk(); _r2.withdraw()
            _mb.showerror(
                "Install failed",
                f"Could not auto-install: {', '.join(pkgs)}\n\n"
                "Please run manually in a Command Prompt:\n\n"
                f"    py -m pip install {' '.join(pkgs)}",
            )
            _r2.destroy()
        except Exception:
            pass

# Skip auto-install in frozen .exe (PyInstaller) — packages are already bundled
if not getattr(sys, "frozen", False):
    _auto_install_packages()


import csv
import datetime as dt
import hashlib
import os
import re
import shutil
import sqlite3
import struct
import subprocess
import tempfile
import threading
import time
import traceback
import zipfile
import zlib
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import xml.etree.ElementTree as ET
import base64

try:
    import tkinter as tk
    from theme_manager import ThemeManager, apply_theme_to_window, get_copyright_year
    from tkinter import filedialog, messagebox, ttk, simpledialog
except Exception as exc:
    raise RuntimeError("Tkinter is required. Use the standard Windows Python installer.") from exc

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
except Exception:
    Image = None
    ImageDraw = None
    ImageFont = None
    ImageTk = None

try:
    import openpyxl
except Exception as exc:
    raise RuntimeError("openpyxl is required. Install with: py -m pip install openpyxl") from exc

APP_NAME = "GFH Telecom LLC Inventory Audit"
EMBEDDED_LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_logo_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_logo_b64.txt"), "r").read().strip()
EMBEDDED_ICON_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_icon_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_icon_b64.txt"), "r").read().strip()

if getattr(sys, "frozen", False):
    PACKAGE_DIR = Path(sys.executable).resolve().parent
else:
    PACKAGE_DIR = Path(__file__).resolve().parent
PORTABLE_APP_DIR = PACKAGE_DIR / "GFH_Inventory_Audit_Data"
LEGACY_APP_DIR = Path.home() / "GFH_Inventory_Variance_GUI"

# Always use the portable folder next to the script so every laptop sharing
# the same network/USB location opens the SAME database file.
# A per-user fallback would silently create per-machine copies — removed.
def _choose_app_dir() -> Path:
    try:
        PORTABLE_APP_DIR.mkdir(parents=True, exist_ok=True)
        test_file = PORTABLE_APP_DIR / ".write_test"
        test_file.write_text("ok", encoding="utf-8")
        try:
            test_file.unlink()
        except Exception:
            pass
        return PORTABLE_APP_DIR
    except Exception as e:
        # Cannot write next to the script — tell the user explicitly instead of
        # silently creating a per-machine copy that goes out of sync.
        import tkinter as _tk, tkinter.messagebox as _mb
        _root = _tk.Tk(); _root.withdraw()
        _mb.showerror(
            "Folder not writable",
            f"Cannot write to the data folder next to the script:\n"
            f"  {PORTABLE_APP_DIR}\n\n"
            f"Error: {e}\n\n"
            f"Please run the script from a writable location (e.g. a shared "
            f"network folder or USB drive) so all laptops share the same database.",
        )
        _root.destroy()
        raise SystemExit(1)

APP_DIR = _choose_app_dir()
DB_PATH = APP_DIR / "inventory_variance_log.sqlite3"
IMAGE_DIR = APP_DIR / "whatsapp_images"
EXPORT_DIR = APP_DIR / "exports"
STORE_CONFIG_PATH = APP_DIR / "store_master_list.csv"
PACKAGED_STORE_LIST_PATH = PACKAGE_DIR / "store_master_list.csv"


def _resource_path(name: str) -> Path:
    """Resolve a bundled asset for a PyInstaller onefile EXE.

    onefile builds extract data files to sys._MEIPASS at runtime; fall back
    to the directory next to the app/script for normal runs."""
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            bundled = Path(meipass) / name
            if bundled.exists():
                return bundled
    return PACKAGE_DIR / name


HEADER_LOGO_PATH = _resource_path("GFH_Telecom_Logo.png")
STATUS_LOGO_PATH = PACKAGE_DIR / "gfh_telecom_llc_logo.png"
APP_ICON_PATH = PACKAGE_DIR / "gfh_telecom_llc_icon.ico"
APP_ICON_PNG_PATH = PACKAGE_DIR / "gfh_telecom_llc_logo.png"

# ---------------------------------------------------------------------------
# Embedded icon assets — written to PACKAGE_DIR at startup so no external
# file is needed. Swap the base64 strings to change the window icon.
# ---------------------------------------------------------------------------
APP_ICON_ICO_BASE64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "app_icon_ico_base64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "app_icon_ico_base64.txt"), "r").read().strip()

APP_ICON_PNG_BASE64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "app_icon_png_base64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "app_icon_png_base64.txt"), "r").read().strip()


def ensure_app_icon_files() -> None:
    """Decode and write embedded ICO + PNG icons to PACKAGE_DIR once."""
    try:
        ico_bytes = base64.b64decode(APP_ICON_ICO_BASE64.strip().replace("\n", ""))
        if not APP_ICON_PATH.exists() or APP_ICON_PATH.stat().st_size != len(ico_bytes):
            APP_ICON_PATH.write_bytes(ico_bytes)
    except Exception:
        pass
    try:
        png_bytes = base64.b64decode(APP_ICON_PNG_BASE64.strip().replace("\n", ""))
        if not APP_ICON_PNG_PATH.exists() or APP_ICON_PNG_PATH.stat().st_size != len(png_bytes):
            APP_ICON_PNG_PATH.write_bytes(png_bytes)
    except Exception:
        pass



SEARCH_SHORTCUT = "ctrl+f"
SEND_ONLY_UNSENT_BY_DEFAULT = True
AUTO_SEND_CHUNK_SIZE = 22

EXCEL_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
R_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

SEND_MODE_LABELS = {
    "district": "District",
    "store": "Store",
    "rep": "Sales Rep",
}



DEFAULT_STORE_MASTER = []

def default_store_master_records() -> List[Dict[str, str]]:
    return [{"District": district, "Store": store} for district, store in DEFAULT_STORE_MASTER]


def migrate_legacy_app_data_if_needed() -> None:
    """No-op: legacy per-user folder migration removed.
    All instances now share the single portable DB next to the script."""
    pass


def save_store_master_records(records: List[Dict[str, str]], path: Path = STORE_CONFIG_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    normalized_rows: Dict[str, Dict[str, str]] = {}
    for rec in records:
        district = normalize_district(rec.get("District", ""))
        store = display_store(rec.get("Store", ""))
        norm = normalize_store(store)
        if not norm or not district or district == "Unknown":
            continue
        normalized_rows[norm] = {"District": district, "Store": store}
    ordered = sorted(normalized_rows.values(), key=lambda r: (normalize_district(r["District"]), r["Store"].lower()))
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["District", "Store"])
        writer.writeheader()
        writer.writerows(ordered)


def ensure_store_master_file() -> None:
    if STORE_CONFIG_PATH.exists():
        return
    STORE_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    if PACKAGED_STORE_LIST_PATH.exists():
        shutil.copy2(PACKAGED_STORE_LIST_PATH, STORE_CONFIG_PATH)
    else:
        save_store_master_records(default_store_master_records(), STORE_CONFIG_PATH)


def load_store_master_records(path: Path = STORE_CONFIG_PATH) -> List[Dict[str, str]]:
    ensure_store_master_file()
    rows: List[Dict[str, str]] = []
    try:
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for rec in reader:
                district = normalize_district(rec.get("District", ""))
                store = display_store(rec.get("Store", ""))
                if district and district != "Unknown" and store:
                    rows.append({"District": district, "Store": store})
    except Exception:
        rows = default_store_master_records()
        save_store_master_records(rows, path)
    if not rows:
        rows = default_store_master_records()
        save_store_master_records(rows, path)
    return rows


@dataclass
class VarianceRow:
    key: str
    district: str
    store: str
    product: str
    imei: str
    status: str
    created_by: str = ""
    rep_name: str = ""
    created_date: str = ""
    document_status: str = ""
    source_file: str = ""
    cleared: bool = False
    sent_count: int = 0
    last_sent_at: str = ""
    cleared_at: str = ""
    notes: str = ""


@dataclass
class InventoryStatusRow:
    key: str
    district: str
    store: str
    status: str
    rep_name: str = ""
    source_file: str = ""

def now_text() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", safe_text(text).lower())


def normalize_store(text: str) -> str:
    value = safe_text(text)
    value = re.sub(r"\s+", " ", value).strip()
    value = value.replace("–", "-").replace("—", "-")
    return value.lower()


def display_store(text: str) -> str:
    value = safe_text(text)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_phone(text: str) -> str:
    value = safe_text(text)
    if not value:
        return ""
    value = value.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    value = value.replace(".", "")
    if value.startswith("00"):
        value = "+" + value[2:]
    return value


def whatsapp_mention(phone: str) -> str:
    phone = normalize_phone(phone)
    if not phone:
        return ""
    if phone.startswith("@"):
        return phone
    return "@" + phone


def person_name_key(text: str) -> str:
    value = safe_text(text).lower()
    value = value.replace(",", " ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    parts = [p for p in value.split() if p]
    return " ".join(sorted(parts))


def device_rule_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", safe_text(text).lower())


def device_matches_rule(product: str, imei: str, rule_text: str, match_type: str) -> bool:
    rule_clean = device_rule_key(rule_text)
    if not rule_clean:
        return False

    product_clean = device_rule_key(product)
    imei_clean = device_rule_key(imei)
    match_type_clean = normalize_header(match_type)

    if match_type_clean in {"productexact", "exactproduct"}:
        return product_clean == rule_clean
    if match_type_clean in {"imeiexact", "serialexact", "esnexact"}:
        return imei_clean == rule_clean
    if match_type_clean in {"anycontains", "containsany"}:
        return rule_clean in product_clean or rule_clean in imei_clean
    return rule_clean in product_clean


def normalize_district(text: str) -> str:
    value = safe_text(text).replace("–", "-").replace("—", "-")
    value = re.sub(r"\s+", " ", value).strip()
    if not value:
        return "Unknown"
    lower = value.lower()
    if lower.startswith("arizona"):
        return "Arizona"
    aliases = {
        "colorado west": "Colorado West",
        "houston": "Houston",
        "colorado east": "Colorado East",
        "tennessee": "Tennessee",
        "louisiana": "Louisiana",
    }
    return aliases.get(lower, value)


def group_name_for_district(district: str, db=None) -> str:
    normalized = normalize_district(district)
    if db is not None:
        try:
            saved = db.find_whatsapp_group(normalized)
            if saved:
                return saved
        except Exception:
            pass
    return f"GFH TELECOM {normalized.upper()}"


def is_sim_product(product: str) -> bool:
    text = safe_text(product).lower()
    compact = normalize_header(product)
    if not text and not compact:
        return False
    if re.search(r"(^|[^a-z0-9])e?[-\s]?sim(s)?([^a-z0-9]|$)", text):
        return True
    if "simcard" in compact or compact in {"sim", "sims", "esim", "esims"}:
        return True
    if compact.startswith(("sim", "esim")) and any(token in compact for token in ("card", "kit", "pack", "starter")):
        return True
    return False


def excel_serial_to_datetime(value) -> Optional[dt.datetime]:
    text = safe_text(value)
    if not text:
        return None
    try:
        serial = float(text)
    except ValueError:
        # Try common date strings if present.
        for fmt in ("%m/%d/%Y", "%m/%d/%Y %I:%M %p", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return dt.datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None
    base = dt.datetime(1899, 12, 30)
    return base + dt.timedelta(days=serial)


def excel_serial_to_date_text(value) -> str:
    parsed = excel_serial_to_datetime(value)
    if parsed is None:
        return safe_text(value)
    return parsed.strftime("%m/%d/%Y %I:%M %p")


def numeric_excel_date(value) -> float:
    text = safe_text(value)
    try:
        return float(text)
    except ValueError:
        parsed = excel_serial_to_datetime(text)
        if parsed is None:
            return -1.0
        base = dt.datetime(1899, 12, 30)
        return (parsed - base).total_seconds() / 86400.0


def variance_key(store: str, imei: str, product: str, status: str, created_by: str = "", created_date: str = "") -> str:
    raw = "|".join([
        normalize_store(store),
        safe_text(imei).lower(),
        safe_text(product).lower(),
        safe_text(status).lower(),
        safe_text(created_by).lower(),
        safe_text(created_date).lower(),
    ])
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()


class RobustXlsxReader:
    """Small XLSX reader with a fallback for exports missing a central ZIP directory."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.parts = self._read_parts()
        self.shared_strings = self._read_shared_strings()

    def _read_parts(self) -> Dict[str, bytes]:
        parts: Dict[str, bytes] = {}
        try:
            with zipfile.ZipFile(self.path) as zf:
                for name in zf.namelist():
                    parts[name] = zf.read(name)
            return parts
        except Exception:
            data = self.path.read_bytes()
            pos = 0
            while pos + 30 <= len(data) and data[pos:pos + 4] == b"PK\x03\x04":
                try:
                    (
                        _sig,
                        _ver,
                        _flag,
                        method,
                        _mtime,
                        _mdate,
                        _crc,
                        csize,
                        _usize,
                        nlen,
                        xlen,
                    ) = struct.unpack_from("<IHHHHHIIIHH", data, pos)
                    name_start = pos + 30
                    name_end = name_start + nlen
                    name = data[name_start:name_end].decode("utf-8", errors="replace")
                    start = name_end + xlen
                    comp = data[start:min(start + csize, len(data))]
                    if method == 8:
                        try:
                            content = zlib.decompress(comp, -15)
                        except Exception:
                            content = b""
                    elif method == 0:
                        content = comp
                    else:
                        content = b""
                    if content:
                        parts[name] = content
                    pos = start + csize
                except Exception:
                    break
            if not parts:
                raise RuntimeError(f"Could not read Excel file: {self.path}")
            return parts

    @staticmethod
    def _clean_xml(raw: bytes) -> bytes:
        return raw.lstrip(b"\xef\xbb\xbf")

    def _read_shared_strings(self) -> List[str]:
        raw = self.parts.get("xl/sharedStrings.xml")
        if not raw:
            return []
        root = ET.fromstring(self._clean_xml(raw))
        strings: List[str] = []
        for si in root.findall(EXCEL_NS + "si"):
            strings.append("".join((t.text or "") for t in si.iter(EXCEL_NS + "t")))
        return strings

    def _workbook_rels(self) -> Dict[str, str]:
        raw = self.parts.get("xl/_rels/workbook.xml.rels")
        if not raw:
            return {}
        root = ET.fromstring(self._clean_xml(raw))
        rels: Dict[str, str] = {}
        for rel in root.findall(REL_NS + "Relationship"):
            rid = rel.attrib.get("Id", "")
            target = rel.attrib.get("Target", "")
            if not target:
                continue
            if not target.startswith("/"):
                target = "xl/" + target
            else:
                target = target.lstrip("/")
            rels[rid] = target
        return rels

    def sheet_paths(self) -> List[Tuple[str, str]]:
        raw = self.parts.get("xl/workbook.xml")
        if not raw:
            if "xl/worksheets/sheet1.xml" in self.parts:
                return [("Sheet1", "xl/worksheets/sheet1.xml")]
            return []
        root = ET.fromstring(self._clean_xml(raw))
        rels = self._workbook_rels()
        sheets: List[Tuple[str, str]] = []
        for sheet in root.findall(".//" + EXCEL_NS + "sheet"):
            name = sheet.attrib.get("name", "Sheet")
            rid = sheet.attrib.get(R_NS + "id", "")
            path = rels.get(rid)
            if path and path in self.parts:
                sheets.append((name, path))
        if not sheets and "xl/worksheets/sheet1.xml" in self.parts:
            sheets.append(("Sheet1", "xl/worksheets/sheet1.xml"))
        return sheets

    @staticmethod
    def _cell_to_col(ref: str) -> Optional[int]:
        match = re.match(r"([A-Z]+)(\d+)", ref or "")
        if not match:
            return None
        col = 0
        for ch in match.group(1):
            col = col * 26 + ord(ch) - 64
        return col

    def _cell_value(self, cell: ET.Element) -> str:
        cell_type = cell.attrib.get("t")
        value_node = cell.find(EXCEL_NS + "v")
        if cell_type == "s":
            if value_node is None or value_node.text is None:
                return ""
            try:
                idx = int(value_node.text)
            except ValueError:
                return ""
            if 0 <= idx < len(self.shared_strings):
                return self.shared_strings[idx]
            return ""
        if cell_type == "inlineStr":
            inline = cell.find(EXCEL_NS + "is")
            if inline is None:
                return ""
            return "".join((t.text or "") for t in inline.iter(EXCEL_NS + "t"))
        if value_node is None or value_node.text is None:
            return ""
        return value_node.text

    def read_sheet(self, preferred_name: Optional[str] = None) -> List[Dict[str, str]]:
        sheets = self.sheet_paths()
        if not sheets:
            raise RuntimeError(f"No worksheet found in {self.path.name}")
        selected_name, selected_path = sheets[0]
        if preferred_name:
            want = preferred_name.lower().strip()
            for name, path in sheets:
                if want in name.lower().strip():
                    selected_name, selected_path = name, path
                    break
        raw = self.parts[selected_path]
        root = ET.fromstring(self._clean_xml(raw))
        rows: List[Tuple[int, Dict[int, str]]] = []
        for row in root.findall(".//" + EXCEL_NS + "row"):
            row_num = int(row.attrib.get("r", "0") or 0)
            values: Dict[int, str] = {}
            for cell in row.findall(EXCEL_NS + "c"):
                col = self._cell_to_col(cell.attrib.get("r", ""))
                if not col:
                    continue
                values[col] = self._cell_value(cell)
            if values:
                rows.append((row_num, values))
        if not rows:
            return []

        header_row_idx, header_values = rows[0]
        max_col = max(header_values)
        headers = [safe_text(header_values.get(i, "")) for i in range(1, max_col + 1)]
        seen: Dict[str, int] = {}
        clean_headers = []
        for i, header in enumerate(headers, start=1):
            name = header if header else f"Column{i}"
            base = name
            if base in seen:
                seen[base] += 1
                name = f"{base}_{seen[base]}"
            else:
                seen[base] = 1
            clean_headers.append(name)

        records: List[Dict[str, str]] = []
        for row_num, value_map in rows[1:]:
            if row_num <= header_row_idx:
                continue
            record: Dict[str, str] = {}
            max_data_col = max(max_col, max(value_map.keys()) if value_map else max_col)
            for i in range(1, max_data_col + 1):
                header = clean_headers[i - 1] if i <= len(clean_headers) else f"Column{i}"
                record[header] = safe_text(value_map.get(i, ""))
            if any(safe_text(v) for v in record.values()):
                records.append(record)
        return records


def find_column(record: Dict[str, str], candidates: Iterable[str]) -> Optional[str]:
    normalized = {normalize_header(k): k for k in record.keys()}
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized:
            return normalized[key]
    for candidate in candidates:
        key = normalize_header(candidate)
        for norm, original in normalized.items():
            if key and (key in norm or norm in key):
                return original
    return None


def read_xlsx_records(path: str | Path) -> List[Dict[str, str]]:
    reader = RobustXlsxReader(path)
    return reader.read_sheet()


def build_store_maps(time_sheet_records: List[Dict[str, str]]) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    district_by_store: Dict[str, str] = {}
    display_by_store: Dict[str, str] = {}
    rep_by_store: Dict[str, str] = {}
    latest_clock_by_store: Dict[str, float] = {}

    if not time_sheet_records:
        return district_by_store, display_by_store, rep_by_store

    sample = time_sheet_records[0]
    store_col = find_column(sample, ["Store"])
    district_col = find_column(sample, ["District"])
    rep_col = find_column(sample, ["Salesperson", "Sales Person", "Rep Name", "Employee", "Employee Name"])
    clock_in_col = find_column(sample, ["Clock In", "Clock-In", "Date", "Work Date"])
    user_login_col = find_column(sample, ["User Login", "Username", "Login"])

    if not store_col:
        return district_by_store, display_by_store, rep_by_store

    for index, rec in enumerate(time_sheet_records):
        store_raw = rec.get(store_col, "")
        norm = normalize_store(store_raw)
        if not norm:
            continue

        display_by_store[norm] = display_store(store_raw)
        if district_col:
            district = normalize_district(rec.get(district_col, ""))
            if district and district != "Unknown":
                district_by_store[norm] = district

        rep_name = safe_text(rec.get(rep_col, "")) if rep_col else ""
        if not rep_name and user_login_col:
            rep_name = safe_text(rec.get(user_login_col, ""))

        date_score = numeric_excel_date(rec.get(clock_in_col, "")) if clock_in_col else float(index)
        # Keep the latest available rep per store. This mirrors the latest-clock-in mapping used by the GFH status file.
        if rep_name and date_score >= latest_clock_by_store.get(norm, -1.0):
            rep_by_store[norm] = rep_name
            latest_clock_by_store[norm] = date_score

    return district_by_store, display_by_store, rep_by_store


def filter_latest_inventory_records(inventory_records: List[Dict[str, str]]) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    """Keep only the latest Inventory_Count_Result_Details rows per Store + Created By.

    GFH exports can include more than one count from the same Created By user for the same store.
    Older rows should not drive the current variance UI because a later recount can convert earlier
    Deficit rows into Matched rows. The Created Date value includes date and time as an Excel serial.
    """
    metrics = {
        "raw_inventory_rows": len(inventory_records),
        "latest_inventory_rows": len(inventory_records),
        "stale_inventory_rows": 0,
        "latest_created_by_groups": 0,
    }
    if not inventory_records:
        return inventory_records, metrics

    sample = inventory_records[0]
    store_col = find_column(sample, ["Store"])
    created_by_col = find_column(sample, ["Created By", "Count By", "User Login"])
    created_date_col = find_column(sample, ["Created Date", "Created Date/Time", "Date Time", "Date"])

    if not store_col or not created_by_col or not created_date_col:
        return inventory_records, metrics

    latest_by_group: Dict[Tuple[str, str], float] = {}
    scores: List[Tuple[Tuple[str, str], float]] = []

    for index, rec in enumerate(inventory_records):
        store_key = normalize_store(rec.get(store_col, "")) or "__unknown_store__"
        created_by_key = safe_text(rec.get(created_by_col, "")).lower().strip() or "__unknown_created_by__"
        group_key = (store_key, created_by_key)
        score = numeric_excel_date(rec.get(created_date_col, ""))
        if score < 0:
            score = float(index) / 1000000.0
        scores.append((group_key, score))
        if group_key not in latest_by_group or score > latest_by_group[group_key]:
            latest_by_group[group_key] = score

    filtered: List[Dict[str, str]] = []
    for rec, (group_key, score) in zip(inventory_records, scores):
        latest_score = latest_by_group.get(group_key, score)
        if abs(score - latest_score) <= 0.0000001:
            filtered.append(rec)

    metrics["latest_inventory_rows"] = len(filtered)
    metrics["stale_inventory_rows"] = len(inventory_records) - len(filtered)
    metrics["latest_created_by_groups"] = len(latest_by_group)
    return filtered, metrics


def extract_variances(
    inventory_records: List[Dict[str, str]],
    time_sheet_records: List[Dict[str, str]],
    master_store_records: Optional[List[Dict[str, str]]] = None,
    source_file: str = "",
) -> Tuple[List[VarianceRow], Dict[str, int]]:
    if not inventory_records:
        return [], {"completed": 0, "pending": 0, "stores_total": 0, "skipped_sims": 0, "raw_inventory_rows": 0, "latest_inventory_rows": 0, "stale_inventory_rows": 0, "latest_created_by_groups": 0}

    district_by_store, display_by_store, rep_by_store = build_store_maps(time_sheet_records)
    for rec in master_store_records or []:
        district = normalize_district(rec.get("District", ""))
        store = display_store(rec.get("Store", ""))
        norm = normalize_store(store)
        if not norm:
            continue
        if district and district != "Unknown":
            district_by_store[norm] = district
        if store and norm not in display_by_store:
            display_by_store[norm] = store
    sample = inventory_records[0]
    store_col = find_column(sample, ["Store"])
    product_col = find_column(sample, ["Product Description", "Product"])
    imei_col = find_column(sample, ["Serial #", "Serial", "IMEI", "ESN"])
    status_col = find_column(sample, ["Status"])
    created_by_col = find_column(sample, ["Created By", "Count By", "User Login"])
    created_date_col = find_column(sample, ["Created Date", "Date"])
    document_status_col = find_column(sample, ["Document Status"])

    missing = [
        name for name, col in [
            ("Store", store_col),
            ("Product Description", product_col),
            ("Serial # / IMEI", imei_col),
            ("Status", status_col),
        ] if not col
    ]
    if missing:
        raise RuntimeError("Missing required inventory column(s): " + ", ".join(missing))

    inventory_records, latest_metrics = filter_latest_inventory_records(inventory_records)

    # Deduplicate by (Store, IMEI) to prevent duplicate IMEI rows
    dedup_map: Dict[Tuple[str, str], Tuple[Dict[str, str], float]] = {}
    for rec in inventory_records:
        store_raw = rec.get(store_col, "")
        imei = safe_text(rec.get(imei_col, ""))
        if not imei:
            continue
        norm_store = normalize_store(store_raw)
        key = (norm_store, imei.lower())
        date_score = numeric_excel_date(rec.get(created_date_col, "")) if created_date_col else 0
        if key not in dedup_map or date_score > dedup_map[key][1]:
            dedup_map[key] = (rec, date_score)
    inventory_records = [rec for rec, _ in dedup_map.values()]
    latest_metrics["dedup_inventory_rows"] = len(inventory_records)

    completed_store_norms = set()
    for rec in inventory_records:
        norm = normalize_store(rec.get(store_col, ""))
        if norm:
            completed_store_norms.add(norm)

    all_store_norms = set(display_by_store.keys()) | completed_store_norms
    pending_store_norms = all_store_norms - completed_store_norms

    variance_rows: List[VarianceRow] = []
    skipped_sims = 0

    for rec in inventory_records:
        status = safe_text(rec.get(status_col, ""))
        if not status:
            continue
        if normalize_header(status) in {"matched", "match", "ok", "balanced"}:
            continue

        store = display_store(rec.get(store_col, ""))
        product = safe_text(rec.get(product_col, ""))
        imei = safe_text(rec.get(imei_col, ""))

        if not imei:
            continue
        if product and normalize_header(product) in {"accessorycommission", "accessoriescommission"}:
            continue
        if is_sim_product(product):
            skipped_sims += 1
            continue

        norm_store = normalize_store(store)
        district = normalize_district(district_by_store.get(norm_store, "Unknown"))
        rep_name = safe_text(rep_by_store.get(norm_store, ""))
        created_by = safe_text(rec.get(created_by_col, "")) if created_by_col else ""
        created_date = excel_serial_to_date_text(rec.get(created_date_col, "")) if created_date_col else ""
        document_status = safe_text(rec.get(document_status_col, "")) if document_status_col else ""
        key = variance_key(store, imei, product, status, created_by, created_date)

        variance_rows.append(
            VarianceRow(
                key=key,
                district=district,
                store=store,
                product=product,
                imei=imei,
                status=status,
                created_by=created_by,
                rep_name=rep_name,
                created_date=created_date,
                document_status=document_status,
                source_file=source_file,
            )
        )

    summary = {
        "completed": len(completed_store_norms),
        "pending": len(pending_store_norms),
        "stores_total": len(all_store_norms),
        "skipped_sims": skipped_sims,
        **latest_metrics,
    }
    return variance_rows, summary




def build_inventory_status_rows(
    inventory_records: List[Dict[str, str]],
    time_sheet_records: List[Dict[str, str]],
    master_store_records: Optional[List[Dict[str, str]]] = None,
    source_file: str = "",
) -> Tuple[List[InventoryStatusRow], Dict[str, int]]:
    district_by_store, display_by_store, rep_by_store = build_store_maps(time_sheet_records)
    master_display_by_store: Dict[str, str] = {}
    for rec in master_store_records or []:
        district = normalize_district(rec.get("District", ""))
        store = display_store(rec.get("Store", ""))
        norm = normalize_store(store)
        if not norm:
            continue
        if district and district != "Unknown":
            district_by_store[norm] = district
        if store:
            display_by_store[norm] = store
            master_display_by_store[norm] = store
    inv_display_by_store: Dict[str, str] = {}
    completed_store_norms: set[str] = set()

    filtered_records = inventory_records
    latest_metrics = {
        "raw_inventory_rows": len(inventory_records),
        "latest_inventory_rows": len(inventory_records),
        "stale_inventory_rows": 0,
        "latest_created_by_groups": 0,
    }

    if inventory_records:
        sample = inventory_records[0]
        store_col = find_column(sample, ["Store"])
        if store_col:
            filtered_records, latest_metrics = filter_latest_inventory_records(inventory_records)
            for rec in filtered_records:
                store_raw = rec.get(store_col, "")
                norm = normalize_store(store_raw)
                if not norm:
                    continue
                inv_display_by_store[norm] = display_store(store_raw)
                completed_store_norms.add(norm)

    all_store_norms = sorted(set(master_display_by_store.keys()) | set(display_by_store.keys()) | set(inv_display_by_store.keys()) | completed_store_norms)
    rows: List[InventoryStatusRow] = []
    for norm in all_store_norms:
        store = master_display_by_store.get(norm) or display_by_store.get(norm) or inv_display_by_store.get(norm) or norm.title()
        district = normalize_district(district_by_store.get(norm, "Unknown"))
        rep_name = safe_text(rep_by_store.get(norm, ""))
        status = "Completed" if norm in completed_store_norms else "Pending"
        key_raw = f"{norm}|{district}|{status}|status"
        key = hashlib.sha1(key_raw.encode("utf-8", errors="ignore")).hexdigest()
        rows.append(InventoryStatusRow(key=key, district=district, store=store, status=status, rep_name=rep_name, source_file=source_file))

    summary = {
        "completed": len(completed_store_norms),
        "pending": len([r for r in rows if r.status == "Pending"]),
        "stores_total": len(rows),
        **latest_metrics,
    }
    return rows, summary


class VarianceDatabase:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    def connect(self):
        con = sqlite3.connect(self.db_path, timeout=60)
        try:
            con.execute("PRAGMA busy_timeout=60000")
        except Exception:
            pass
        return con

    @staticmethod
    def _table_columns(con: sqlite3.Connection, table: str) -> set[str]:
        try:
            return {row[1] for row in con.execute(f"PRAGMA table_info({table})").fetchall()}
        except Exception:
            return set()

    def _init_db(self) -> None:
        with self.connect() as con:
            try:
                con.execute("PRAGMA journal_mode=WAL")
                con.execute("PRAGMA synchronous=NORMAL")
            except Exception:
                pass
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS variances (
                    key TEXT PRIMARY KEY,
                    district TEXT,
                    store TEXT,
                    product TEXT,
                    imei TEXT,
                    status TEXT,
                    created_by TEXT,
                    rep_name TEXT,
                    created_date TEXT,
                    document_status TEXT,
                    source_file TEXT,
                    first_seen_at TEXT,
                    last_seen_at TEXT,
                    cleared INTEGER DEFAULT 0,
                    cleared_at TEXT,
                    sent_count INTEGER DEFAULT 0,
                    last_sent_at TEXT,
                    notes TEXT
                )
                """
            )
            variance_columns = self._table_columns(con, "variances")
            for column, definition in {
                "rep_name": "TEXT DEFAULT ''",
                "created_by": "TEXT DEFAULT ''",
                "created_date": "TEXT DEFAULT ''",
                "document_status": "TEXT DEFAULT ''",
                "source_file": "TEXT DEFAULT ''",
                "notes": "TEXT DEFAULT ''",
            }.items():
                if column not in variance_columns:
                    con.execute(f"ALTER TABLE variances ADD COLUMN {column} {definition}")

            con.execute(
                """
                CREATE TABLE IF NOT EXISTS send_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    variance_key TEXT,
                    district TEXT,
                    group_name TEXT,
                    batch_title TEXT,
                    image_path TEXT,
                    sent_at TEXT,
                    mode TEXT,
                    error TEXT
                )
                """
            )
            send_columns = self._table_columns(con, "send_log")
            if "batch_title" not in send_columns:
                con.execute("ALTER TABLE send_log ADD COLUMN batch_title TEXT DEFAULT ''")

            con.execute(
                """
                CREATE TABLE IF NOT EXISTS inventory_status_state (
                    store_key TEXT PRIMARY KEY,
                    district TEXT,
                    store TEXT,
                    last_status TEXT,
                    previous_status TEXT,
                    status_changed_at TEXT,
                    last_sent_at TEXT,
                    last_loaded_at TEXT
                )
                """
            )

            con.execute(
                """
                CREATE TABLE IF NOT EXISTS whatsapp_groups (
                    district TEXT PRIMARY KEY,
                    group_name TEXT NOT NULL,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )

            con.execute(
                """
                CREATE TABLE IF NOT EXISTS district_managers (
                    district TEXT PRIMARY KEY,
                    dm_name TEXT,
                    phone TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )

            con.execute(
                """
                CREATE TABLE IF NOT EXISTS store_accounts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    district TEXT NOT NULL,
                    store TEXT NOT NULL,
                    store_key TEXT UNIQUE NOT NULL,
                    account_id TEXT,
                    username TEXT,
                    password TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS sales_reps (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    rep_name TEXT NOT NULL,
                    rep_key TEXT UNIQUE NOT NULL,
                    phone TEXT NOT NULL,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )

            con.execute(
                """
                CREATE TABLE IF NOT EXISTS device_exclusions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    rule_text TEXT NOT NULL,
                    rule_key TEXT UNIQUE NOT NULL,
                    match_type TEXT NOT NULL,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
            exclusion_columns = self._table_columns(con, "device_exclusions")
            if "product" not in exclusion_columns:
                con.execute("ALTER TABLE device_exclusions ADD COLUMN product TEXT DEFAULT ''")
            if "imei" not in exclusion_columns:
                con.execute("ALTER TABLE device_exclusions ADD COLUMN imei TEXT DEFAULT ''")
            if "comments" not in exclusion_columns:
                con.execute("ALTER TABLE device_exclusions ADD COLUMN comments TEXT DEFAULT ''")
            if "district" not in exclusion_columns:
                con.execute("ALTER TABLE device_exclusions ADD COLUMN district TEXT DEFAULT ''")

    def upsert_rows(self, rows: List[VarianceRow]) -> None:
        with self.connect() as con:
            for row in rows:
                existing = con.execute(
                    "SELECT cleared, cleared_at, sent_count, last_sent_at, notes, first_seen_at FROM variances WHERE key=?",
                    (row.key,),
                ).fetchone()
                if existing:
                    con.execute(
                        """
                        UPDATE variances
                        SET district=?, store=?, product=?, imei=?, status=?, created_by=?, rep_name=?, created_date=?,
                            document_status=?, source_file=?, last_seen_at=?
                        WHERE key=?
                        """,
                        (
                            row.district, row.store, row.product, row.imei, row.status,
                            row.created_by, row.rep_name, row.created_date, row.document_status,
                            row.source_file, now_text(), row.key,
                        ),
                    )
                else:
                    con.execute(
                        """
                        INSERT INTO variances (
                            key, district, store, product, imei, status, created_by, rep_name, created_date,
                            document_status, source_file, first_seen_at, last_seen_at, cleared,
                            cleared_at, sent_count, last_sent_at, notes
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, '', 0, '', '')
                        """,
                        (
                            row.key, row.district, row.store, row.product, row.imei, row.status,
                            row.created_by, row.rep_name, row.created_date, row.document_status, row.source_file,
                            now_text(), now_text(),
                        ),
                    )

    def rows(self, include_cleared: bool = False) -> List[VarianceRow]:
        query = """
            SELECT key, district, store, product, imei, status, created_by, rep_name, created_date,
                   document_status, source_file, cleared, sent_count, last_sent_at, cleared_at, notes
            FROM variances
        """
        if not include_cleared:
            query += " WHERE cleared=0"
        query += " ORDER BY district, store, rep_name, status, product"
        with self.connect() as con:
            data = con.execute(query).fetchall()
        rows: List[VarianceRow] = []
        for item in data:
            rows.append(
                VarianceRow(
                    key=item[0], district=item[1] or "", store=item[2] or "", product=item[3] or "",
                    imei=item[4] or "", status=item[5] or "", created_by=item[6] or "",
                    rep_name=item[7] or "", created_date=item[8] or "", document_status=item[9] or "",
                    source_file=item[10] or "", cleared=bool(item[11]), sent_count=int(item[12] or 0),
                    last_sent_at=item[13] or "", cleared_at=item[14] or "", notes=item[15] or "",
                )
            )
        return rows

    def get_rows_by_keys(self, keys: Iterable[str]) -> List[VarianceRow]:
        keys = list(keys)
        if not keys:
            return []
        placeholders = ",".join("?" for _ in keys)
        with self.connect() as con:
            data = con.execute(
                f"""
                SELECT key, district, store, product, imei, status, created_by, rep_name, created_date,
                       document_status, source_file, cleared, sent_count, last_sent_at, cleared_at, notes
                FROM variances WHERE key IN ({placeholders})
                ORDER BY district, store, rep_name, status, product
                """,
                keys,
            ).fetchall()
        return [
            VarianceRow(
                key=i[0], district=i[1] or "", store=i[2] or "", product=i[3] or "", imei=i[4] or "",
                status=i[5] or "", created_by=i[6] or "", rep_name=i[7] or "", created_date=i[8] or "",
                document_status=i[9] or "", source_file=i[10] or "", cleared=bool(i[11]),
                sent_count=int(i[12] or 0), last_sent_at=i[13] or "", cleared_at=i[14] or "", notes=i[15] or "",
            )
            for i in data
        ]

    def set_cleared(self, key: str, cleared: bool) -> None:
        with self.connect() as con:
            con.execute(
                "UPDATE variances SET cleared=?, cleared_at=? WHERE key=?",
                (1 if cleared else 0, now_text() if cleared else "", key),
            )

    def mark_sent(self, rows: Iterable[VarianceRow], group_name: str, batch_title: str, image_path: str, mode: str, error: str = "") -> None:
        sent_at = now_text()
        with self.connect() as con:
            for row in rows:
                con.execute(
                    """
                    INSERT INTO send_log (variance_key, district, group_name, batch_title, image_path, sent_at, mode, error)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (row.key, row.district, group_name, batch_title, image_path, sent_at, mode, error),
                )
                if not error:
                    con.execute(
                        "UPDATE variances SET sent_count=COALESCE(sent_count,0)+1, last_sent_at=? WHERE key=?",
                        (sent_at, row.key),
                    )


    def upsert_inventory_status_rows(self, rows: List[InventoryStatusRow]) -> None:
        loaded_at = now_text()
        with self.connect() as con:
            for row in rows:
                store_key = normalize_store(row.store)
                if not store_key:
                    continue
                existing = con.execute(
                    "SELECT last_status, previous_status, status_changed_at, last_sent_at FROM inventory_status_state WHERE store_key=?",
                    (store_key,),
                ).fetchone()

                last_status = safe_text(row.status)
                if existing:
                    old_status = safe_text(existing[0])
                    old_previous = safe_text(existing[1])
                    old_changed_at = safe_text(existing[2])
                    old_sent_at = safe_text(existing[3])
                    if old_status and normalize_header(old_status) != normalize_header(last_status):
                        previous_status = old_status
                        changed_at = loaded_at
                        # New status means new status still needs its own send state.
                        sent_at = ""
                    else:
                        previous_status = old_previous
                        changed_at = old_changed_at
                        sent_at = old_sent_at
                    con.execute(
                        """
                        UPDATE inventory_status_state
                        SET district=?, store=?, last_status=?, previous_status=?, status_changed_at=?, last_sent_at=?, last_loaded_at=?
                        WHERE store_key=?
                        """,
                        (row.district, row.store, last_status, previous_status, changed_at, sent_at, loaded_at, store_key),
                    )
                else:
                    con.execute(
                        """
                        INSERT INTO inventory_status_state
                        (store_key, district, store, last_status, previous_status, status_changed_at, last_sent_at, last_loaded_at)
                        VALUES (?, ?, ?, ?, '', '', '', ?)
                        """,
                        (store_key, row.district, row.store, last_status, loaded_at),
                    )

    def mark_status_sent(self, rows: Iterable[InventoryStatusRow]) -> None:
        sent_at = now_text()
        with self.connect() as con:
            for row in rows:
                store_key = normalize_store(row.store)
                if not store_key:
                    continue
                con.execute(
                    """
                    INSERT INTO inventory_status_state
                    (store_key, district, store, last_status, previous_status, status_changed_at, last_sent_at, last_loaded_at)
                    VALUES (?, ?, ?, ?, '', '', ?, ?)
                    ON CONFLICT(store_key) DO UPDATE SET
                        district=excluded.district,
                        store=excluded.store,
                        last_status=excluded.last_status,
                        last_sent_at=excluded.last_sent_at,
                        last_loaded_at=excluded.last_loaded_at
                    """,
                    (store_key, row.district, row.store, row.status, sent_at, sent_at),
                )

    def inventory_status_state_map(self) -> Dict[str, Dict[str, str]]:
        with self.connect() as con:
            rows = con.execute(
                """
                SELECT store_key, district, store, last_status, previous_status, status_changed_at, last_sent_at, last_loaded_at
                FROM inventory_status_state
                """
            ).fetchall()
        return {
            r[0]: {
                "District": r[1] or "",
                "Store": r[2] or "",
                "Last Status": r[3] or "",
                "Previous Status": r[4] or "",
                "Status Changed At": r[5] or "",
                "Last Sent At": r[6] or "",
                "Last Loaded At": r[7] or "",
            }
            for r in rows
        }

    def save_whatsapp_group(self, district: str, group_name: str) -> None:
        district = normalize_district(district)
        group_name = safe_text(group_name).strip()
        if not district or district == "Unknown":
            raise ValueError("District is required.")
        if not group_name:
            raise ValueError("WhatsApp group name is required.")
        with self.connect() as con:
            con.execute(
                """
                INSERT INTO whatsapp_groups (district, group_name, created_at, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(district) DO UPDATE SET
                    group_name=excluded.group_name,
                    updated_at=excluded.updated_at
                """,
                (district, group_name, now_text(), now_text()),
            )

    def delete_whatsapp_group(self, district: str) -> None:
        with self.connect() as con:
            con.execute("DELETE FROM whatsapp_groups WHERE district=?", (normalize_district(district),))

    def whatsapp_groups(self) -> List[Dict[str, str]]:
        with self.connect() as con:
            rows = con.execute(
                "SELECT district, group_name FROM whatsapp_groups ORDER BY district"
            ).fetchall()
        return [{"District": r[0] or "", "Group Name": r[1] or ""} for r in rows]

    def find_whatsapp_group(self, district: str) -> str:
        with self.connect() as con:
            row = con.execute(
                "SELECT group_name FROM whatsapp_groups WHERE district=?",
                (normalize_district(district),),
            ).fetchone()
        return safe_text(row[0]) if row else ""

    def save_district_manager(self, district: str, dm_name: str, phone: str) -> None:
        district = normalize_district(district)
        dm_name = safe_text(dm_name)
        phone = normalize_phone(phone)
        if not district or district == "Unknown":
            raise ValueError("District is required.")
        if not phone:
            raise ValueError("DM phone number is required.")
        with self.connect() as con:
            con.execute(
                """
                INSERT INTO district_managers (district, dm_name, phone, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(district) DO UPDATE SET
                    dm_name=excluded.dm_name,
                    phone=excluded.phone,
                    updated_at=excluded.updated_at
                """,
                (district, dm_name, phone, now_text(), now_text()),
            )

    def delete_district_manager(self, district: str) -> None:
        with self.connect() as con:
            con.execute("DELETE FROM district_managers WHERE district=?", (normalize_district(district),))

    def district_managers(self) -> List[Dict[str, str]]:
        with self.connect() as con:
            rows = con.execute(
                "SELECT district, dm_name, phone FROM district_managers ORDER BY district"
            ).fetchall()
        return [{"District": r[0] or "", "DM Name": r[1] or "", "Phone": r[2] or ""} for r in rows]

    def all_known_districts(self) -> List[str]:
        """Return sorted list of unique districts from all district-having tables."""
        districts = set()
        with self.connect() as con:
            for table in ("store_accounts", "whatsapp_groups", "district_managers", "device_exclusions", "variance_log"):
                try:
                    rows = con.execute(f"SELECT DISTINCT district FROM {table} WHERE district IS NOT NULL AND district != ''").fetchall()
                    for r in rows:
                        d = safe_text(r[0]).strip()
                        if d:
                            districts.add(normalize_district(d))
                except Exception:
                    pass
        return sorted(districts)

    def find_district_manager_phone(self, district: str) -> str:
        with self.connect() as con:
            row = con.execute(
                "SELECT phone FROM district_managers WHERE district=?",
                (normalize_district(district),),
            ).fetchone()
        return normalize_phone(row[0]) if row else ""

    def save_store_account(self, district: str, store: str, account_id: str, username: str, password: str) -> None:
        district = normalize_district(district)
        store = display_store(store)
        store_key = normalize_store(store)
        if not store_key:
            raise ValueError("Store name is required.")
        with self.connect() as con:
            con.execute(
                """
                INSERT INTO store_accounts (district, store, store_key, account_id, username, password, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(store_key) DO UPDATE SET
                    district=excluded.district,
                    store=excluded.store,
                    account_id=excluded.account_id,
                    username=excluded.username,
                    password=excluded.password,
                    updated_at=excluded.updated_at
                """,
                (district, store, store_key, safe_text(account_id), safe_text(username), safe_text(password), now_text(), now_text()),
            )

    def delete_store_account(self, store_key: str) -> None:
        with self.connect() as con:
            con.execute("DELETE FROM store_accounts WHERE store_key=?", (store_key,))

    def store_accounts(self) -> List[Dict[str, str]]:
        with self.connect() as con:
            rows = con.execute(
                """
                SELECT district, store, store_key, account_id, username, password
                FROM store_accounts
                ORDER BY district, store
                """
            ).fetchall()
        return [
            {
                "District": row[0] or "",
                "Store": row[1] or "",
                "StoreKey": row[2] or "",
                "Account ID": row[3] or "",
                "Username": row[4] or "",
                "Password": row[5] or "",
            }
            for row in rows
        ]

    def replace_store_accounts(self, records: List[Dict[str, str]]) -> int:
        cleaned: Dict[str, Dict[str, str]] = {}
        for rec in records:
            district = normalize_district(rec.get("District", ""))
            store = display_store(rec.get("Store", ""))
            store_key = normalize_store(store)
            if not store_key or not district or district == "Unknown":
                continue
            cleaned[store_key] = {"District": district, "Store": store, "StoreKey": store_key}

        now = now_text()
        last_error = None
        for attempt in range(5):
            try:
                with self.connect() as con:
                    con.execute("BEGIN IMMEDIATE")
                    con.execute("DELETE FROM store_accounts")
                    for rec in sorted(cleaned.values(), key=lambda r: (r["District"], r["Store"].lower())):
                        con.execute(
                            """
                            INSERT INTO store_accounts (district, store, store_key, account_id, username, password, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (rec["District"], rec["Store"], rec["StoreKey"], "", "", "", now, now),
                        )
                    con.commit()
                return len(cleaned)
            except sqlite3.OperationalError as exc:
                last_error = exc
                if "locked" not in str(exc).lower():
                    raise
                time.sleep(0.8 + attempt * 0.6)
        raise RuntimeError("Database is locked. Close any other running copy of this GUI, then import again.") from last_error

    def store_master_records(self) -> List[Dict[str, str]]:
        return [{"District": row["District"], "Store": row["Store"]} for row in self.store_accounts()]

    def save_sales_rep(self, rep_name: str, phone: str) -> None:
        rep_name = safe_text(rep_name)
        phone = normalize_phone(phone)
        rep_key = person_name_key(rep_name)
        if not rep_key:
            raise ValueError("Sales rep name is required.")
        if not phone:
            raise ValueError("Phone number is required.")
        with self.connect() as con:
            con.execute(
                """
                INSERT INTO sales_reps (rep_name, rep_key, phone, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(rep_key) DO UPDATE SET
                    rep_name=excluded.rep_name,
                    phone=excluded.phone,
                    updated_at=excluded.updated_at
                """,
                (rep_name, rep_key, phone, now_text(), now_text()),
            )

    def delete_sales_rep(self, rep_key: str) -> None:
        with self.connect() as con:
            con.execute("DELETE FROM sales_reps WHERE rep_key=?", (rep_key,))

    def sales_reps(self) -> List[Dict[str, str]]:
        with self.connect() as con:
            rows = con.execute(
                "SELECT rep_name, rep_key, phone FROM sales_reps ORDER BY rep_name"
            ).fetchall()
        return [{"Rep Name": r[0] or "", "RepKey": r[1] or "", "Phone": r[2] or ""} for r in rows]

    def sales_rep_phone_map(self) -> Dict[str, str]:
        return {row["RepKey"]: row["Phone"] for row in self.sales_reps()}

    def find_sales_rep_phone(self, rep_name: str) -> str:
        key = person_name_key(rep_name)
        if not key:
            return ""
        reps = self.sales_reps()
        exact = {row["RepKey"]: row["Phone"] for row in reps}
        if key in exact:
            return exact[key]
        key_parts = set(key.split())
        for row in reps:
            rep_parts = set((row["RepKey"] or "").split())
            if key_parts and rep_parts and (key_parts <= rep_parts or rep_parts <= key_parts):
                return row["Phone"]
        return ""

    def update_device_exclusion_comment(self, rule_key: str, comments: str) -> None:
        with self.connect() as con:
            con.execute(
                "UPDATE device_exclusions SET comments=?, updated_at=? WHERE rule_key=?",
                (safe_text(comments), now_text(), rule_key),
            )

    def save_device_exclusion(self, product: str = "", imei: str = "", comments: str = "", district: str = "") -> None:
        product = safe_text(product)
        imei = safe_text(imei)
        comments = safe_text(comments)
        district = normalize_district(district) if safe_text(district) else ""

        if not product and not imei:
            raise ValueError("Product or IMEI is required.")

        # IMEI exclusions stay global across all districts.
        if imei:
            district = ""
            rule_text = imei
            match_type = "IMEI Exact"
        else:
            if not district or district == "Unknown":
                raise ValueError("District is required for Product exclusions.")
            rule_text = product
            match_type = "Product Contains"

        # Include district in product exclusion key, but keep IMEI exclusions global.
        rule_key = device_rule_key(f"{district}|{product}|{imei}" if not imei else f"|{imei}")
        if not rule_key:
            raise ValueError("Product or IMEI is required.")

        with self.connect() as con:
            # Prevent duplicates by ensuring old keys or un-districted versions are cleaned up
            if imei:
                con.execute("DELETE FROM device_exclusions WHERE imei=?", (imei,))
            elif product and district:
                con.execute("DELETE FROM device_exclusions WHERE product=? AND district=?", (product, district))

            con.execute(
                """
                INSERT INTO device_exclusions (rule_text, rule_key, match_type, product, imei, comments, district, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(rule_key) DO UPDATE SET
                    rule_text=excluded.rule_text,
                    match_type=excluded.match_type,
                    product=excluded.product,
                    imei=excluded.imei,
                    comments=excluded.comments,
                    district=excluded.district,
                    updated_at=excluded.updated_at
                """,
                (rule_text, rule_key, match_type, product, imei, comments, district, now_text(), now_text()),
            )

    def delete_device_exclusion(self, rule_key: str) -> None:
        with self.connect() as con:
            con.execute("DELETE FROM device_exclusions WHERE rule_key=?", (rule_key,))

    def delete_all_device_exclusions(self) -> None:
        with self.connect() as con:
            con.execute("DELETE FROM device_exclusions")

    def device_exclusions(self) -> List[Dict[str, str]]:
        with self.connect() as con:
            rows = con.execute(
                """
                SELECT rule_text, rule_key, match_type, product, imei, comments, district
                FROM device_exclusions
                ORDER BY COALESCE(district, ''), COALESCE(product, ''), COALESCE(imei, '')
                """
            ).fetchall()

        out: List[Dict[str, str]] = []
        for r in rows:
            rule_text = r[0] or ""
            rule_key = r[1] or ""
            match_type = r[2] or ""
            product = r[3] or ""
            imei = r[4] or ""
            comments = r[5] or ""
            district = normalize_district(r[6] or "") if safe_text(r[6] or "") else ""

            # Backfill display values for exclusions created in old versions.
            if not product and not imei and rule_text:
                if normalize_header(match_type) in {"imeiexact", "serialexact", "esnexact"}:
                    imei = rule_text
                    district = ""
                else:
                    product = rule_text

            if imei:
                district = ""

            out.append({
                "District": district,
                "Product": product,
                "IMEI": imei,
                "Comments": comments,
                "Rule Text": rule_text,
                "RuleKey": rule_key,
                "Match Type": match_type,
            })
        return out

    def is_device_excluded(self, district: str, product: str, imei: str) -> bool:
        row_district = normalize_district(district)
        product_clean = device_rule_key(product)
        imei_clean = device_rule_key(imei)

        for rule in self.device_exclusions():
            rule_product = device_rule_key(rule.get("Product", ""))
            rule_imei = device_rule_key(rule.get("IMEI", ""))
            rule_district = normalize_district(rule.get("District", "")) if safe_text(rule.get("District", "")) else ""

            # IMEI exclusions stay global. District is ignored for IMEI.
            if rule_imei and imei_clean and rule_imei == imei_clean:
                return True

            # Product exclusions are district-specific only.
            if rule_product and product_clean and rule_product in product_clean:
                if rule_district and normalize_district(rule_district) == row_district:
                    return True

        return False

    def exclusion_reason(self, district: str, product: str, imei: str) -> str:
        row_district = normalize_district(district)
        product_clean = device_rule_key(product)
        imei_clean = device_rule_key(imei)

        for rule in self.device_exclusions():
            rule_product = device_rule_key(rule.get("Product", ""))
            rule_imei = device_rule_key(rule.get("IMEI", ""))
            rule_district = normalize_district(rule.get("District", "")) if safe_text(rule.get("District", "")) else ""

            if rule_imei and imei_clean and rule_imei == imei_clean:
                return f"IMEI: {rule.get('IMEI', '')}"

            if rule_product and product_clean and rule_product in product_clean:
                if rule_district and normalize_district(rule_district) == row_district:
                    return f"{rule_district} Product: {rule.get('Product', '')}"

        return ""

    def export_xlsx(self, output_path: Path, keys: Optional[Iterable[str]] = None) -> None:
        keys = list(keys) if keys is not None else None
        where_sql = ""
        params: List[str] = []
        if keys is not None:
            if not keys:
                headers = [
                    "District", "Store", "Product", "IMEI", "Status", "Rep Name", "Count By", "Created Date",
                    "Document Status", "Clearance Status", "Cleared At", "Sent Count", "Last Sent At",
                    "First Seen At", "Last Seen At", "Source File", "Notes"
                ]
                output_path.parent.mkdir(parents=True, exist_ok=True)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)
                wb.save(output_path)
                return
            placeholders = ",".join("?" for _ in keys)
            where_sql = f"WHERE key IN ({placeholders})"
            params = keys

        query = f"""
                SELECT district, store, product, imei, status, rep_name, created_by, created_date,
                       document_status,
                       CASE WHEN cleared=1 THEN 'Cleared' ELSE 'Not Cleared' END,
                       cleared_at, sent_count, last_sent_at, first_seen_at, last_seen_at, source_file, notes
                FROM variances
                {where_sql}
                ORDER BY district, store, rep_name, status, product
                """
        with self.connect() as con:
            rows = con.execute(query, params).fetchall()

        headers = [
            "District", "Store", "Product", "IMEI", "Status", "Rep Name", "Count By", "Created Date",
            "Document Status", "Clearance Status", "Cleared At", "Sent Count", "Last Sent At",
            "First Seen At", "Last Seen At", "Source File", "Notes"
        ]

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)

        imei_col_idx = headers.index("IMEI") + 1

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == imei_col_idx:
                    cell.data_type = 's'
                    cell.number_format = '@'

        wb.save(output_path)

class ImageRenderer:
    def __init__(self):
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        if Image is None:
            raise RuntimeError("Pillow is required. Install with: py -m pip install pillow")

    @staticmethod
    def _font(size: int, bold: bool = False):
        candidates = []
        if bold:
            candidates.extend(["arialbd.ttf", "Arial Bold.ttf", "segoeuib.ttf"])
        candidates.extend(["arial.ttf", "Segoe UI.ttf", "DejaVuSans.ttf"])
        for candidate in candidates:
            try:
                return ImageFont.truetype(candidate, size)
            except Exception:
                continue
        return ImageFont.load_default()

    @staticmethod
    def _wrap(draw, text: str, font, max_width: int) -> List[str]:
        text = safe_text(text)
        if not text:
            return [""]
        words = text.split()
        lines: List[str] = []
        current = ""
        for word in words:
            probe = word if not current else current + " " + word
            bbox = draw.textbbox((0, 0), probe, font=font)
            if bbox[2] - bbox[0] <= max_width:
                current = probe
            else:
                if current:
                    lines.append(current)
                current = word
        if current:
            lines.append(current)
        return lines or [""]

    def render_rows(self, batch_title: str, rows: List[VarianceRow], mode: str = "pending") -> Path:
        if not rows:
            raise ValueError("No rows to render")

        width = 1560
        margin = 32
        title_font = self._font(34, True)
        sub_font = self._font(18, False)
        header_font = self._font(20, True)
        cell_font = self._font(18, False)
        small_font = self._font(16, False)
        bold_small_font = self._font(16, True)
        row_height_base = 52

        tmp = Image.new("RGB", (width, 400), "white")
        draw = ImageDraw.Draw(tmp)
        col_widths = [165, 220, 480, 205, 140, 250]
        product_width = col_widths[2] - 20
        rep_width = col_widths[5] - 20
        row_heights = []
        for row in rows:
            product_lines = self._wrap(draw, str(row.product or ""), cell_font, product_width)
            rep_lines = self._wrap(draw, str(row.rep_name or ""), small_font, rep_width)
            row_heights.append(max(row_height_base, 26 * max(len(product_lines), len(rep_lines)) + 22))

        dark = (18, 20, 43)
        gray = (95, 95, 102)
        red = (233, 27, 47)
        light = (246, 247, 249)
        border = (215, 218, 223)

        logo_img = None
        logo_w = 0
        logo_h_used = 0
        if STATUS_LOGO_PATH.exists():
            try:
                logo_img = Image.open(STATUS_LOGO_PATH).convert("RGBA")
                scale = min(640 / logo_img.width, 150 / logo_img.height)
                size = (max(1, int(logo_img.width * scale)), max(1, int(logo_img.height * scale)))
                logo_img = logo_img.resize(size)
                logo_w, logo_h_used = size
            except Exception:
                logo_img = None
                logo_w = 0
                logo_h_used = 0

        header_area_h = max(logo_h_used, 74) + 18
        table_top = margin + header_area_h + 18
        table_height = 54 + sum(row_heights)
        height = table_top + table_height + 58

        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)

        clean_mode = safe_text(mode).replace("manual_", "").replace("pending_", "")
        mode_label = SEND_MODE_LABELS.get(clean_mode, clean_mode.replace("_", " ").title())

        draw.text((margin, margin + 4), "GFH Inventory Variance", fill=dark, font=title_font)
        draw.text(
            (margin, margin + 48),
            f"{mode_label}: {batch_title}   Rows: {len(rows)}   Generated: {now_text()}",
            fill=gray,
            font=sub_font,
        )

        if logo_img is not None:
            logo_x = width - margin - logo_w
            logo_y = margin
            img.paste(logo_img, (logo_x, logo_y), logo_img)

        y = table_top
        x = margin
        headers = ["District", "Store", "Product", "IMEI", "Status", "Rep Name"]
        draw.rectangle((x, y, width - margin, y + 54), fill=red)
        cx = x
        for idx, header in enumerate(headers):
            draw.text((cx + 10, y + 14), header, fill="white", font=header_font)
            cx += col_widths[idx]
        y += 54

        for idx, row in enumerate(rows):
            fill = light if idx % 2 == 0 else (255, 255, 255)
            rh = row_heights[idx]
            draw.rectangle((x, y, width - margin, y + rh), fill=fill, outline=border)
            values = [row.district, row.store, row.product, row.imei, row.status, row.rep_name]
            cx = x
            for cidx, value in enumerate(values):
                max_w = col_widths[cidx] - 20
                font = bold_small_font if cidx in {0, 1, 5} else (small_font if cidx == 4 else cell_font)
                lines = self._wrap(draw, str(value or ""), font, max_w)
                yy = y + 11
                for line in lines[:4]:
                    draw.text((cx + 10, yy), line, fill=dark, font=font)
                    yy += 25
                cx += col_widths[cidx]
            y += rh

        footer = "Please provide resolution image or valid variance explanation. Cleared variances will not be auto-sent again."
        draw.text((margin, height - 36), footer, fill=gray, font=small_font)

        safe_title = re.sub(r"[^A-Za-z0-9_-]+", "_", batch_title)[:60] or "Variance"
        safe_mode = re.sub(r"[^A-Za-z0-9_-]+", "_", safe_text(mode).replace("manual_", "").replace("pending_", ""))[:30] or "mode"
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = IMAGE_DIR / f"GFH_Variance_{safe_mode}_{safe_title}_{stamp}.png"
        img.save(path)
        return path


class WhatsAppSender:
    def __init__(self, status_callback=None):
        self.status_callback = status_callback or (lambda text: None)

    def log(self, text: str) -> None:
        self.status_callback(text)

    @staticmethod
    def _import_pyautogui():
        try:
            import pyautogui
            pyautogui.FAILSAFE = True
            return pyautogui
        except Exception as exc:
            raise RuntimeError("pyautogui is required. Install with: py -m pip install pyautogui") from exc

    @staticmethod
    def _open_whatsapp() -> None:
        if sys.platform.startswith("win"):
            try:
                subprocess.Popen("start whatsapp:", shell=True)
                time.sleep(3)
                return
            except Exception:
                pass
        time.sleep(1)

    @staticmethod
    def _activate_whatsapp_window() -> bool:
        try:
            import pygetwindow as gw
            windows = [w for w in gw.getAllWindows() if "whatsapp" in (w.title or "").lower()]
            if not windows:
                return False
            win = windows[0]
            if win.isMinimized:
                win.restore()
            win.activate()
            time.sleep(1)
            return True
        except Exception:
            return False

    @staticmethod
    def _copy_image_to_clipboard(image_path: Path) -> None:
        if Image is None:
            raise RuntimeError("Pillow is required to copy images to clipboard.")
        if not sys.platform.startswith("win"):
            raise RuntimeError("Automatic image clipboard paste is implemented for Windows only.")
        try:
            import win32clipboard
            import win32con
        except Exception as exc:
            raise RuntimeError("pywin32 is required. Install with: py -m pip install pywin32") from exc
        image = Image.open(image_path).convert("RGB")
        output = BytesIO()
        image.save(output, "BMP")
        data = output.getvalue()[14:]
        output.close()
        win32clipboard.OpenClipboard()
        try:
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32con.CF_DIB, data)
        finally:
            win32clipboard.CloseClipboard()

    @staticmethod
    def _type_group_search(pyautogui, group_name: str) -> None:
        shortcut = SEARCH_SHORTCUT.lower().strip()
        if "+" in shortcut:
            keys = [k.strip() for k in shortcut.split("+") if k.strip()]
        else:
            keys = [k.strip() for k in shortcut.split() if k.strip()]
        if not keys:
            keys = ["ctrl", "f"]
        pyautogui.hotkey(*keys)
        time.sleep(0.7)
        pyautogui.hotkey("ctrl", "a")
        time.sleep(0.2)
        pyautogui.write(group_name, interval=0.01)
        time.sleep(1.2)
        pyautogui.press("enter")
        time.sleep(1.5)

    @staticmethod
    def _paste_text(pyautogui, text: str) -> None:
        text = safe_text(text)
        if not text:
            return
        try:
            import pyperclip
            pyperclip.copy(text)
            pyautogui.hotkey("ctrl", "v")
        except Exception:
            pyautogui.write(text, interval=0.01)

    def send_image(self, group_name: str, image_path: Path, text_message: str = "") -> None:
        pyautogui = self._import_pyautogui()
        self.log(f"Opening WhatsApp Desktop for {group_name}...")
        self._open_whatsapp()
        self._activate_whatsapp_window()
        self.log(f"Searching group: {group_name}")
        self._type_group_search(pyautogui, group_name)

        self.log("Copying image to clipboard...")
        self._copy_image_to_clipboard(image_path)

        # Paste the image — this opens WhatsApp's image-preview/send dialog
        pyautogui.hotkey("ctrl", "v")
        # Wait for the preview dialog to fully load before doing anything else
        time.sleep(3.0)

        caption = safe_text(text_message)
        if caption:
            self.log("Typing caption into image preview field...")
            # WhatsApp Desktop focuses the caption field automatically when the
            # image preview dialog opens. Type directly into it.
            self._paste_text(pyautogui, caption)
            time.sleep(0.8)

        # Send the image (Enter confirms the image-preview dialog)
        pyautogui.press("enter")
        time.sleep(2.0)
        self.log(f"Sent image to {group_name}")


    def send_text(self, group_name: str, text_message: str) -> None:
        message = safe_text(text_message)
        if not message:
            return
        pyautogui = self._import_pyautogui()
        self.log(f"Opening WhatsApp Desktop for {group_name}...")
        self._open_whatsapp()
        self._activate_whatsapp_window()
        self.log(f"Searching group: {group_name}")
        self._type_group_search(pyautogui, group_name)
        self.log("Sending WhatsApp text message...")
        self._paste_text(pyautogui, message)
        time.sleep(0.7)
        pyautogui.press("enter")
        time.sleep(1.0)
        self.log(f"Sent text to {group_name}")




def show_startup_error(exc: BaseException) -> None:
    error_text = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    try:
        APP_DIR.mkdir(parents=True, exist_ok=True)
        (APP_DIR / "startup_error.log").write_text(error_text, encoding="utf-8")
    except Exception:
        try:
            Path(tempfile.gettempdir(), "gfh_inventory_audit_startup_error.log").write_text(error_text, encoding="utf-8")
        except Exception:
            pass
    try:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("GFH Inventory Audit Error", str(exc))
        root.destroy()
    except Exception:
        print("GFH Inventory Audit Error:", exc)
        print(error_text)


GFH_SQUARE_ICON_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "gfh_square_icon_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "gfh_square_icon_b64.txt"), "r").read().strip()


class GFHApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self._app_icon = None
        # Dynamic screen resolution support: size to 90% of the screen and
        # center it (DPI-aware), then stay a normal resizable top-level so
        # Windows Snap (50% left/right, corners, Win+arrow) keeps working.
        self._apply_dynamic_geometry()
        # Try _MEIPASS first (PyInstaller onefile extraction dir)
        import sys as _sys, os as _os
        _meipass = getattr(_sys, "_MEIPASS", None)
        if _meipass:
            for _ico_name in ("gfh_icon.ico", "gfh_telecom_llc_icon.ico", "gfh_icon.ico"):
                _ico_path = _os.path.join(_meipass, _ico_name)
                if _os.path.exists(_ico_path):
                    try:
                        self.iconbitmap(_ico_path)
                        self.iconbitmap(_ico_path)
                    except Exception:
                        pass
                    break
        # Fallback: decode EMBEDDED_ICON_B64 to %TEMP%
        try:
            import base64 as _b64, tempfile as _tf
            _data = _b64.b64decode(EMBEDDED_ICON_B64.strip())
            _tmp_dir = _os.environ.get("TEMP", _tf.gettempdir())
            _ico_path = _os.path.join(_tmp_dir, "gfh_audit_icon.ico")
            with open(_ico_path, "wb") as _f:
                _f.write(_data)
            self.iconbitmap(_ico_path)
            self.iconbitmap(_ico_path)
        except Exception:
            pass
        self.zoom_scale = 1.0
        APP_DIR.mkdir(parents=True, exist_ok=True)
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        migrate_legacy_app_data_if_needed()
        self.db = VarianceDatabase(DB_PATH)
        # Track file mtime for cross-instance sync.
        try:
            self._db_mtime: float = DB_PATH.stat().st_mtime if DB_PATH.exists() else 0.0
        except Exception:
            self._db_mtime = 0.0
        self._db_sync_paused: bool = False   # paused while THIS instance is writing
        self.master_store_records = self.db.store_master_records()
        self.current_inventory_records: List[Dict[str, str]] = []
        self.current_time_sheet_records: List[Dict[str, str]] = []

        self.store_district_var = tk.StringVar(value="")
        self.store_name_var = tk.StringVar(value="")
        self.rep_name_var = tk.StringVar(value="")
        self.rep_phone_var = tk.StringVar(value="")
        self.dm_district_var = tk.StringVar(value="")
        self.dm_name_var = tk.StringVar(value="")
        self.dm_phone_var = tk.StringVar(value="")
        self.wg_district_var = tk.StringVar(value="")
        self.wg_group_name_var = tk.StringVar(value="")
        self.store_search_var = tk.StringVar(value="")
        self.rep_search_var = tk.StringVar(value="")
        self.exclusion_search_var = tk.StringVar(value="")
        self.exclusion_district_var = tk.StringVar(value="")
        self.exclusion_product_var = tk.StringVar(value="")
        self.exclusion_imei_var = tk.StringVar(value="")
        self.exclusion_comments_var = tk.StringVar(value="")
        self.selected_exclusion_key_var = tk.StringVar(value="")

        self.inventory_path = tk.StringVar(value="")
        self.time_sheet_path = tk.StringVar(value="")
        self.status_text = tk.StringVar(value="Select both file locations, then click Load Variances. No data loaded yet.")
        self.summary_text = tk.StringVar(value="No data loaded")
        self.include_cleared = tk.BooleanVar(value=False)
        self.send_only_unsent = tk.BooleanVar(value=SEND_ONLY_UNSENT_BY_DEFAULT)

        self.status_send_mode = tk.StringVar(value="District")
        self.status_district_filter = tk.StringVar(value="All Districts")
        self.status_store_filter = tk.StringVar(value="All Stores")
        self.status_search_any_var = tk.StringVar(value="")
        self.status_search_district_var = tk.StringVar(value="")
        self.status_search_store_var = tk.StringVar(value="")
        self.status_search_status_var = tk.StringVar(value="")
        self.status_search_rep_var = tk.StringVar(value="")

        self.audit_send_mode = tk.StringVar(value="District")
        self.audit_district_filter = tk.StringVar(value="All Districts")
        self.audit_store_filter = tk.StringVar(value="All Stores")
        self.audit_search_any_var = tk.StringVar(value="")
        self.audit_search_district_var = tk.StringVar(value="")
        self.audit_search_store_var = tk.StringVar(value="")
        self.audit_search_product_var = tk.StringVar(value="")
        self.audit_search_imei_var = tk.StringVar(value="")
        self.audit_search_rep_var = tk.StringVar(value="")
        self.final_district_var = tk.StringVar(value="All Districts")

        self.loaded_keys: set[str] = set()
        self.data_loaded = False
        self.key_by_iid: Dict[str, str] = {}

        self.status_rows: List[InventoryStatusRow] = []
        self.status_row_by_key: Dict[str, InventoryStatusRow] = {}
        self.status_key_by_iid: Dict[str, str] = {}
        self.status_checked_keys: set[str] = set()
        self.audit_checked_keys: set[str] = set()

        self.theme_manager = ThemeManager("GFH Inventory Audit", app_name="vidapay-gfh")
        self._build_ui()
        self.set_status(f"Ready. Data folder: {APP_DIR}")
        self._start_db_sync_poll()
        apply_theme_to_window(self, self.theme_manager)

    def _apply_dynamic_geometry(self) -> None:
        """Size the window to 90% of the screen and center it.

        Works on any laptop/monitor/PC (1080p, 1440p, 2K, 4K) and respects
        Windows DPI scaling (run after _enable_dpi_awareness()). The window
        stays resizable so Windows Snap gestures keep working — it centers
        on launch, then snaps normally to 50% left/right, corners or via
        Win+arrow shortcuts.
        """
        try:
            self.update_idletasks()
            sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
            w = max(960, min(int(sw * 0.90), sw - 20))
            h = max(640, min(int(sh * 0.90), sh - 40))
            x = max(0, (sw - w) // 2)
            y = max(0, (sh - h) // 2)
            self.geometry(f"{w}x{h}+{x}+{y}")
            # minsize <= half the screen so 50% / corner snap is never blocked
            self.minsize(min(960, max(640, sw // 2)),
                         min(580, max(480, sh // 2)))
            self.resizable(True, True)
        except Exception:
            pass

    def _apply_styles(self) -> None:
        sz = lambda n: max(6, round(n * self.zoom_scale))
        s = self._style
        s.configure("TFrame", background=self.COLOR_BG)
        s.configure("Card.TFrame", background=self.COLOR_CARD)
        s.configure("Brand.TFrame", background=self.COLOR_NAVY)
        s.configure("TLabel", background=self.COLOR_BG, foreground=self.COLOR_TEXT, font=("Segoe UI", sz(10)))
        s.configure("Header.TLabel", font=("Segoe UI", sz(19), "bold"), background=self.COLOR_NAVY, foreground="#FFFFFF")
        s.configure("BrandSub.TLabel", font=("Segoe UI", sz(10), "bold"), background=self.COLOR_NAVY, foreground="#DCE2F2")
        s.configure("Sub.TLabel", font=("Segoe UI", sz(10)), background=self.COLOR_BG, foreground=self.COLOR_MUTED)
        s.configure("TButton", padding=(10, 6), font=("Segoe UI", sz(9), "bold"), background="#FFFFFF", foreground=self.COLOR_NAVY, bordercolor=self.COLOR_BORDER, focusthickness=1, focuscolor=self.COLOR_RED)
        s.map(
            "TButton",
            background=[("active", "#FFE8EC"), ("pressed", self.COLOR_RED)],
            foreground=[("pressed", "#FFFFFF"), ("active", self.COLOR_NAVY)],
            bordercolor=[("active", self.COLOR_RED), ("pressed", self.COLOR_RED)],
        )
        s.configure("TEntry", fieldbackground="#FFFFFF", foreground=self.COLOR_TEXT, bordercolor=self.COLOR_BORDER)
        s.configure("TCombobox", fieldbackground="#FFFFFF", foreground=self.COLOR_TEXT, bordercolor=self.COLOR_BORDER)
        s.configure("TLabelframe", background=self.COLOR_BG, bordercolor=self.COLOR_BORDER, relief="solid")
        s.configure("TLabelframe.Label", background=self.COLOR_BG, foreground=self.COLOR_NAVY, font=("Segoe UI", sz(10), "bold"))
        s.configure("TNotebook", background=self.COLOR_BG, borderwidth=0)
        s.configure("TNotebook.Tab", padding=(18, 9), font=("Segoe UI", sz(10), "bold"), background="#E9ECF5", foreground=self.COLOR_NAVY)
        s.map(
            "TNotebook.Tab",
            background=[("selected", self.COLOR_RED), ("active", "#FFE8EC")],
            foreground=[("selected", "#FFFFFF"), ("active", self.COLOR_NAVY)],
        )
        s.configure("Treeview", rowheight=max(20, round(32 * self.zoom_scale)), font=("Segoe UI", sz(10)), background="#FFFFFF", fieldbackground="#FFFFFF", foreground=self.COLOR_TEXT, bordercolor=self.COLOR_BORDER, borderwidth=1)
        s.configure("Treeview.Heading", font=("Segoe UI", sz(10), "bold"), background=self.COLOR_NAVY, foreground="#FFFFFF", relief="flat")
        s.map("Treeview", background=[("selected", self.COLOR_RED)], foreground=[("selected", "#FFFFFF")])

    def zoom_in(self, event=None) -> None:
        if self.zoom_scale < 2.0:
            self.zoom_scale = round(self.zoom_scale + 0.1, 1)
            self.apply_zoom()

    def zoom_out(self, event=None) -> None:
        if self.zoom_scale > 0.5:
            self.zoom_scale = round(self.zoom_scale - 0.1, 1)
            self.apply_zoom()

    def apply_zoom(self) -> None:
        self._apply_styles()
        self.update_idletasks()

    def _build_ui(self) -> None:
        self.header_mgr = FixedHeaderManager(self, title="GFH Inventory Audit")
        self.header_mgr.add_theme_toggle(self.theme_manager, callback=self._apply_theme)
        # FixedHeaderManager now tags ALL its own widgets with _tag="header"
        # in __init__/add_theme_toggle/add_copyright, so no manual tagging needed.
        try:
            _lp = _resource_path("GFH_Telecom_Logo.png") if "_resource_path" in dir() else os.path.join(os.path.dirname(os.path.abspath(__file__)), "GFH_Telecom_Logo.png")
            if os.path.exists(_lp):
                self.header_mgr.set_logo(logo_path=_lp, text="GFH")
        except Exception:
            pass
        self._style = ttk.Style(self)
        try:
            self._style.theme_use("clam")
        except Exception:
            pass

        self.COLOR_NAVY = "#090d26"   # matches theme_manager.py navy — header blends with logo
        self.COLOR_RED = "#f0541c"
        self.COLOR_BG = "#F3F5FA"
        self.COLOR_CARD = "#FFFFFF"
        self.COLOR_TEXT = "#090d26"
        self.COLOR_MUTED = "#5F6678"
        self.COLOR_BORDER = "#D9DEEA"
        self.COLOR_SUCCESS = "#17A65B"

        self.configure(bg=self.COLOR_BG)
        self._apply_styles()

        # ── Copyright bar (bottom) ──
        _cbar = tk.Frame(self, bg="#090d26", height=24)
        _cbar.pack(fill="x", side="bottom")
        _cbar.pack_propagate(False)
        tk.Label(
            _cbar,
            text=f"Developed by Abad Umair Channa | Copyright © {date.today().year} | All rights reserved.",
            font=("Segoe UI", 8), fg="#9d9db8", bg="#090d26",
        ).pack(side="left", padx=14, pady=3)


        root = ttk.Frame(self, padding=14)
        root.pack(fill="both", expand=True)

        header = ttk.Frame(root, style="Brand.TFrame", padding=(18, 14))
        header._tag = "header"
        header.pack(fill="x")

        self.header_logo_img = None
        if HEADER_LOGO_PATH.exists() and Image is not None and ImageTk is not None:
            try:
                logo = Image.open(HEADER_LOGO_PATH).convert("RGBA")
                scale = min(560 / logo.width, 116 / logo.height)
                size = (max(1, int(logo.width * scale)), max(1, int(logo.height * scale)))
                logo = logo.resize(size)
                self.header_logo_img = ImageTk.PhotoImage(logo)
                tk.Label(header, image=self.header_logo_img, bg=self.COLOR_NAVY, bd=0).pack(side="left", padx=(0, 18))
            except Exception:
                self.header_logo_img = None

        title_wrap = ttk.Frame(header, style="Brand.TFrame")
        title_wrap.pack(side="left", fill="x", expand=True)
        ttk.Label(title_wrap, text=APP_NAME, style="Header.TLabel").pack(anchor="w")
        ttk.Label(title_wrap, text="Inventory Status • Variance Audit • WhatsApp Dispatch", style="BrandSub.TLabel").pack(anchor="w", pady=(4, 0))
        ttk.Label(title_wrap, textvariable=self.summary_text, style="BrandSub.TLabel").pack(anchor="w", pady=(8, 0))

        file_box = ttk.LabelFrame(root, text="Upload Files", padding=10)
        file_box.pack(fill="x", pady=(12, 8))
        self._file_row(file_box, 0, "Inventory_Count_Result_Details", self.inventory_path, self.pick_inventory)
        self._file_row(file_box, 1, "Employee_Time_Sheet", self.time_sheet_path, self.pick_time_sheet)
        ttk.Button(file_box, text="Load Variances", command=self.load_variances).grid(row=0, column=3, rowspan=2, padx=(12, 0), sticky="ns")
        ttk.Button(file_box, text="−", width=3, command=self.zoom_out).grid(row=0, column=4, padx=(8, 2), sticky="ew")
        ttk.Button(file_box, text="+", width=3, command=self.zoom_in).grid(row=1, column=4, padx=(8, 2), sticky="ew")
        file_box.columnconfigure(1, weight=1)
        self.bind("<Control-equal>", self.zoom_in)
        self.bind("<Control-plus>", self.zoom_in)
        self.bind("<Control-minus>", self.zoom_out)

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, pady=(6, 0))
        self.status_tab = ttk.Frame(self.notebook, padding=10)
        self.audit_tab = ttk.Frame(self.notebook, padding=10)
        self.store_tab = ttk.Frame(self.notebook, padding=10)
        self.rep_tab = ttk.Frame(self.notebook, padding=10)
        self.dm_tab = ttk.Frame(self.notebook, padding=10)
        self.exclusion_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.status_tab, text="Inventory Audit Status")
        self.notebook.add(self.audit_tab, text="Variance Audit")
        self.notebook.add(self.store_tab, text="Store List")
        self.notebook.add(self.rep_tab, text="Employees")
        self.notebook.add(self.dm_tab, text="District DMs")
        self.notebook.add(self.exclusion_tab, text="Excluded Devices")

        self._build_status_tab()
        self._build_audit_tab()
        self._build_store_tab()
        self._build_rep_tab()
        self._build_dm_tab()
        self._build_exclusion_tab()
        self.refresh_store_accounts_table()
        self.refresh_sales_reps_table()
        self.refresh_district_managers_table()
        self.refresh_whatsapp_groups_table()
        self.refresh_device_exclusions_table()

        status_bar = ttk.Label(root, textvariable=self.status_text, anchor="w", relief="sunken", padding=6, foreground=self.COLOR_NAVY, background="#E9ECF5")
        status_bar.pack(fill="x", pady=(8, 0))

        theme_btn = self.theme_manager.create_theme_toggle_button(header, callback=self._apply_theme)
        theme_btn.pack(side="right")

    def _apply_theme(self, colors=None):
        """Apply theme colors to all widgets.

        GFHApp inherits from tk.Tk, so `self` IS the root window.
        Previous code used self.root which doesn't exist → AttributeError
        → theme toggle silently did nothing.
        """
        if colors is None:
            colors = self.theme_manager.get_colors()
        # Pass `self` (the tk.Tk root), not self.root (which doesn't exist)
        self.theme_manager.apply_theme_to_window(self)
        # Refresh header toggle button text in case theme changed
        if hasattr(self.header_mgr, 'update_button_text'):
            self.header_mgr.update_button_text()
    def _build_status_tab(self) -> None:
        controls = ttk.LabelFrame(self.status_tab, text="Send Inventory Audit Status", padding=10)
        controls.pack(fill="x", pady=(0, 5))
        ttk.Label(controls, text="Send by:").grid(row=0, column=0, sticky="w")
        status_mode = ttk.Combobox(controls, textvariable=self.status_send_mode, values=["District", "Store", "Sales Rep"], state="readonly", width=14)
        status_mode.grid(row=0, column=1, padx=(6, 12), sticky="w")
        ttk.Label(controls, text="District filter:").grid(row=0, column=2, sticky="w")
        self.status_district_combo = ttk.Combobox(controls, textvariable=self.status_district_filter, values=["All Districts"], state="readonly", width=22)
        self.status_district_combo.grid(row=0, column=3, padx=(6, 12), sticky="w")
        self.status_district_combo.bind("<<ComboboxSelected>>", self.on_status_district_change)
        ttk.Label(controls, text="Store filter:").grid(row=0, column=4, sticky="w")
        self.status_store_combo = ttk.Combobox(controls, textvariable=self.status_store_filter, values=["All Stores"], state="readonly", width=24)
        self.status_store_combo.grid(row=0, column=5, padx=(6, 12), sticky="w")
        self.status_store_combo.bind("<<ComboboxSelected>>", self.on_status_store_change)
        ttk.Button(controls, text="Check Current Filter", command=self.auto_check_status_rows).grid(row=0, column=6, padx=(0, 6))
        ttk.Button(controls, text="Check Pending Only", command=self.auto_check_pending_status_rows).grid(row=0, column=7, padx=(0, 6))
        ttk.Button(controls, text="Clear Checkmarks", command=self.clear_status_checkmarks).grid(row=0, column=8, padx=(0, 6))
        ttk.Button(controls, text="Send Status Image", command=self.send_checked_status).grid(row=0, column=9, padx=(0, 6))
        ttk.Button(controls, text="Add Store", command=self.add_store_prompt).grid(row=0, column=10, padx=(0, 6))
        ttk.Button(controls, text="Open Folder", command=self.open_app_folder).grid(row=0, column=11, padx=(0, 6))

        # Search — single row
        status_search_box = ttk.LabelFrame(self.status_tab, text="Search Inventory Audit Status", padding=8)
        status_search_box.pack(fill="x", pady=(0, 5))
        status_search_fields = [
            ("Any", self.status_search_any_var, 14),
            ("District", self.status_search_district_var, 12),
            ("Store", self.status_search_store_var, 14),
            ("Status", self.status_search_status_var, 12),
            ("Rep", self.status_search_rep_var, 14),
        ]
        for idx, (label, var, width) in enumerate(status_search_fields):
            ttk.Label(status_search_box, text=label + ":").grid(row=0, column=idx * 2, sticky="w", padx=(0, 2))
            ent = ttk.Entry(status_search_box, textvariable=var, width=width)
            ent.grid(row=0, column=idx * 2 + 1, sticky="w", padx=(0, 6))
            ent.bind("<KeyRelease>", lambda _e: self.refresh_status_table())
        ttk.Button(status_search_box, text="Clear", command=self.clear_status_search).grid(row=0, column=10, sticky="w")

        cols = ("district", "store", "status", "rep_name", "checkbox")
        self.status_tree = ttk.Treeview(self.status_tab, columns=cols, show="headings", selectmode="browse", height=12)
        headings = {
            "district": "District",
            "store": "Store",
            "status": "Status",
            "rep_name": "Rep Name",
            "checkbox": "Checkbox",
        }
        widths = {"district": 130, "store": 200, "status": 110, "rep_name": 170, "checkbox": 95}
        for col in cols:
            self.status_tree.heading(col, text=headings[col], command=lambda c=col: self.sort_any_tree(self.status_tree, c, False))
            self.status_tree.column(col, width=widths[col], minwidth=80, anchor="w")
        self.status_tree.column("checkbox", anchor="center")
        self.status_tree.tag_configure("status_pending", background="#FFF3CD")
        self.status_tree.tag_configure("status_completed_after_update", background="#D7ECFF")
        self.status_tree.tag_configure("status_completed_sent", background="#D9F7DF")
        self.status_tree.tag_configure("status_completed", background="#FFFFFF")
        ttk.Label(
            self.status_tab,
            text="Colors: Yellow = Pending, Blue = Completed after updated sheet load, Green = Completed and sent to WhatsApp.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(0, 4))
        yscroll = ttk.Scrollbar(self.status_tab, orient="vertical", command=self.status_tree.yview)
        xscroll = ttk.Scrollbar(self.status_tab, orient="horizontal", command=self.status_tree.xview)
        self.status_tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        xscroll.pack(side="bottom", fill="x")
        self.status_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        self.status_tree.bind("<Button-1>", self.on_status_tree_click)

    def _build_audit_tab(self) -> None:
        # ── Variance Audit Controls — one scrollable single line ──────────
        # Every control sits in a single row inside a horizontal-scroll
        # canvas, so nothing is ever cut off at small or snapped window
        # sizes: scroll, trackpad drag or Shift+wheel to reach the rest.
        toolbar_host = tk.Frame(self.audit_tab)
        toolbar_host.pack(fill="x", pady=(0, 5))
        toolbar_canvas = tk.Canvas(toolbar_host, highlightthickness=0, bd=0)
        toolbar_hbar = ttk.Scrollbar(toolbar_host, orient="horizontal", command=toolbar_canvas.xview)
        toolbar_canvas.configure(xscrollcommand=toolbar_hbar.set)
        toolbar_hbar.pack(side="bottom", fill="x")
        toolbar_canvas.pack(fill="x")

        send_box = ttk.LabelFrame(toolbar_canvas, text="Variance Audit Controls", padding=10)
        toolbar_canvas.create_window((0, 0), window=send_box, anchor="nw")

        def _sync_toolbar_scroll(_e=None):
            toolbar_canvas.configure(scrollregion=toolbar_canvas.bbox("all"))
            desired = send_box.winfo_reqheight() + 2
            if toolbar_canvas.winfo_height() != desired:
                toolbar_canvas.configure(height=desired)

        def _toolbar_wheel(event):
            if event.state & 0x0001:  # Shift held → page scroll
                factor = 3
            else:
                factor = 1
            toolbar_canvas.xview_scroll(int(-1 * (event.delta / 120)) * factor, "units")
            return "break"

        send_box.bind("<Configure>", _sync_toolbar_scroll)
        toolbar_canvas.bind("<Configure>", _sync_toolbar_scroll)
        toolbar_canvas.bind("<MouseWheel>", _toolbar_wheel)

        ttk.Label(send_box, text="Send by:").grid(row=0, column=0, sticky="w")
        audit_mode_picker = ttk.Combobox(send_box, textvariable=self.audit_send_mode, values=["District", "Store", "Sales Rep"], state="readonly", width=14)
        audit_mode_picker.grid(row=0, column=1, padx=(6, 10), sticky="w")
        ttk.Label(send_box, text="District filter:").grid(row=0, column=2, sticky="w")
        self.audit_district_combo = ttk.Combobox(send_box, textvariable=self.audit_district_filter, values=["All Districts"], state="readonly", width=20)
        self.audit_district_combo.grid(row=0, column=3, padx=(6, 10), sticky="w")
        self.audit_district_combo.bind("<<ComboboxSelected>>", self.on_audit_district_change)
        ttk.Label(send_box, text="Store filter:").grid(row=0, column=4, sticky="w")
        self.audit_store_combo = ttk.Combobox(send_box, textvariable=self.audit_store_filter, values=["All Stores"], state="readonly", width=22)
        self.audit_store_combo.grid(row=0, column=5, padx=(6, 10), sticky="w")
        self.audit_store_combo.bind("<<ComboboxSelected>>", self.on_audit_store_change)
        ttk.Button(send_box, text="Check Current Filter", command=self.auto_check_audit_rows).grid(row=0, column=6, padx=(0, 6))
        ttk.Button(send_box, text="Clear Checkmarks", command=self.clear_audit_checkmarks).grid(row=0, column=7, padx=(0, 6))
        ttk.Button(send_box, text="Send Checked Variance Image", command=self.send_checked_variances).grid(row=0, column=8, padx=(0, 6))
        ttk.Button(send_box, text="Send Selected Image", command=self.send_selected).grid(row=0, column=9, padx=(0, 6))
        ttk.Button(send_box, text="Send Pending", command=self.send_pending).grid(row=0, column=10, padx=(0, 6))
        ttk.Checkbutton(send_box, text="Only unsent", variable=self.send_only_unsent).grid(row=0, column=11, padx=(0, 4), sticky="w")
        ttk.Button(send_box, text="Mark Cleared", command=lambda: self.mark_selected(True)).grid(row=0, column=12, padx=(0, 6))
        ttk.Button(send_box, text="Mark Not Cleared", command=lambda: self.mark_selected(False)).grid(row=0, column=13, padx=(0, 6))
        ttk.Button(send_box, text="Export Log", command=self.export_log).grid(row=0, column=14, padx=(0, 6))
        ttk.Button(send_box, text="Open Folder", command=self.open_app_folder).grid(row=0, column=15, padx=(0, 6))
        ttk.Button(send_box, text="Clear UI", command=self.clear_current_ui).grid(row=0, column=16, padx=(0, 6))
        ttk.Button(send_box, text="Copy IMEI", command=self.copy_selected_imei).grid(row=0, column=17, padx=(0, 6))
        ttk.Checkbutton(send_box, text="Show cleared", variable=self.include_cleared, command=self.refresh_table).grid(row=0, column=18, sticky="w", padx=(4, 0))

        for child in send_box.winfo_children():
            child.bind("<MouseWheel>", _toolbar_wheel)

        # ── Final Send Actions ─────────────────────────────────────────────
        action_box = ttk.LabelFrame(self.audit_tab, text="Final Send Actions", padding=(10, 6))
        action_box.pack(fill="x", pady=(0, 5))
        ttk.Label(action_box, text="District:").grid(row=0, column=0, sticky="w")
        self.final_district_combo = ttk.Combobox(action_box, textvariable=self.final_district_var, values=["All Districts"], state="readonly", width=22)
        self.final_district_combo.grid(row=0, column=1, padx=(6, 12), sticky="w")
        ttk.Button(action_box, text="Send Starting Message", command=self.send_starting_message).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(action_box, text="Send Reminder 1", command=lambda: self.send_reminder(1)).grid(row=0, column=3, padx=(0, 6))
        ttk.Button(action_box, text="Send Reminder 2", command=lambda: self.send_reminder(2)).grid(row=0, column=4, padx=(0, 6))
        ttk.Button(action_box, text="Send Reminder 3", command=lambda: self.send_reminder(3)).grid(row=0, column=5, padx=(0, 6))
        ttk.Button(action_box, text="Send Final District Result", command=self.send_final_district_result).grid(row=0, column=6, padx=(0, 6))

        # Search box — single row
        search_box = ttk.LabelFrame(self.audit_tab, text="Search Variance Rows", padding=8)
        search_box.pack(fill="x", pady=(0, 5))
        search_fields = [
            ("Any", self.audit_search_any_var, 14),
            ("District", self.audit_search_district_var, 12),
            ("Store", self.audit_search_store_var, 14),
            ("Product", self.audit_search_product_var, 16),
            ("IMEI", self.audit_search_imei_var, 14),
            ("Rep", self.audit_search_rep_var, 12),
        ]
        for idx, (label, var, width) in enumerate(search_fields):
            ttk.Label(search_box, text=label + ":").grid(row=0, column=idx * 2, sticky="w", padx=(0, 2))
            ent = ttk.Entry(search_box, textvariable=var, width=width)
            ent.grid(row=0, column=idx * 2 + 1, sticky="w", padx=(0, 6))
            ent.bind("<KeyRelease>", lambda _e: self.refresh_table())
        ttk.Button(search_box, text="Clear", command=self.clear_audit_search).grid(row=0, column=12, sticky="w")

        columns = ("district", "store", "product", "imei", "status", "rep_name", "clearance", "checkbox")
        self.audit_tree = ttk.Treeview(self.audit_tab, columns=columns, show="headings", selectmode="extended", height=12)
        headings = {
            "district": "District",
            "store": "Store",
            "product": "Product",
            "imei": "IMEI",
            "status": "Status",
            "rep_name": "Rep Name",
            "clearance": "Clearance",
            "checkbox": "Checkbox",
        }
        widths = {"district": 110, "store": 150, "product": 280, "imei": 130, "status": 100, "rep_name": 170, "clearance": 110, "checkbox": 85}
        for col in columns:
            self.audit_tree.heading(col, text=headings[col], command=lambda c=col: self.sort_any_tree(self.audit_tree, c, False))
            self.audit_tree.column(col, width=widths[col], minwidth=80, anchor="w")
        self.audit_tree.column("checkbox", anchor="center")
        self.audit_tree.tag_configure("variance_pending", background="#FFF3CD")
        self.audit_tree.tag_configure("variance_sent", background="#D7ECFF")
        self.audit_tree.tag_configure("variance_cleared", background="#D9F7DF")
        ttk.Label(
            self.audit_tab,
            text="Colors: Yellow = Pending variance, Blue = Sent to WhatsApp, Green = Cleared.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(0, 4))
        yscroll = ttk.Scrollbar(self.audit_tab, orient="vertical", command=self.audit_tree.yview)
        xscroll = ttk.Scrollbar(self.audit_tab, orient="horizontal", command=self.audit_tree.xview)
        self.audit_tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        xscroll.pack(side="bottom", fill="x")
        self.audit_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        self.audit_tree.bind("<Double-1>", self.on_audit_double_click)
        self.audit_tree.bind("<Button-1>", self.on_audit_tree_click)
        self.audit_tree.bind("<Control-c>", self.copy_selected_imei)
        self.audit_tree.bind("<Button-3>", self.copy_selected_imei)

    def _build_store_tab(self) -> None:
        form = ttk.LabelFrame(self.store_tab, text="Add or Update Store", padding=10)
        form.pack(fill="x", pady=(0, 8))
        ttk.Label(form, text="District").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=3)
        self._store_district_combo = ttk.Combobox(
            form,
            textvariable=self.store_district_var,
            values=self.db.all_known_districts(),
            width=20,
        )
        self._store_district_combo.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=3)
        ttk.Label(form, text="Store").grid(row=0, column=2, sticky="w", padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=self.store_name_var, width=32).grid(row=0, column=3, sticky="w", padx=(0, 12), pady=3)
        ttk.Button(form, text="Save Store", command=self.save_store_account_from_form).grid(row=0, column=4, padx=(0, 6), pady=3)
        ttk.Button(form, text="Import XLSX", command=self.import_store_accounts_file).grid(row=0, column=5, padx=(0, 6), pady=3)
        ttk.Button(form, text="Delete Selected", command=self.delete_selected_store_account).grid(row=0, column=6, padx=(0, 6), pady=3)
        ttk.Button(form, text="Clear Form", command=self.clear_store_form).grid(row=0, column=7, padx=(0, 6), pady=3)

        # Search bar for Store List
        store_search_box = ttk.Frame(self.store_tab)
        store_search_box.pack(fill="x", pady=(0, 8))
        ttk.Label(store_search_box, text="Search:").pack(side="left", padx=(0, 6))
        store_search_entry = ttk.Entry(store_search_box, textvariable=self.store_search_var, width=40)
        store_search_entry.pack(side="left", padx=(0, 8))
        store_search_entry.bind("<KeyRelease>", lambda _e: self.refresh_store_accounts_table())
        ttk.Button(store_search_box, text="Clear", command=lambda: (self.store_search_var.set(""), self.refresh_store_accounts_table())).pack(side="left")

        columns = ("district", "store")
        self.store_tree = ttk.Treeview(self.store_tab, columns=columns, show="headings", selectmode="browse", height=12)
        headings = {
            "district": "District",
            "store": "Store",
        }
        widths = {"district": 220, "store": 360}
        for col in columns:
            self.store_tree.heading(col, text=headings[col], command=lambda c=col: self.sort_any_tree(self.store_tree, c, False))
            self.store_tree.column(col, width=widths[col], minwidth=120, anchor="w")
        self.store_tree.pack(fill="both", expand=True)
        self.store_tree.bind("<<TreeviewSelect>>", self.on_store_account_select)

    def _build_rep_tab(self) -> None:
        form = ttk.LabelFrame(self.rep_tab, text="Add or Update Employee Phone", padding=10)
        form.pack(fill="x", pady=(0, 8))
        ttk.Label(form, text="Employee Name").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=self.rep_name_var, width=32).grid(row=0, column=1, sticky="w", padx=(0, 12), pady=3)
        ttk.Label(form, text="Phone Number").grid(row=0, column=2, sticky="w", padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=self.rep_phone_var, width=24).grid(row=0, column=3, sticky="w", padx=(0, 12), pady=3)
        ttk.Button(form, text="Save Employee", command=self.save_sales_rep_from_form).grid(row=0, column=4, padx=(0, 6), pady=3)
        ttk.Button(form, text="Import XLSX", command=self.import_employees_file).grid(row=0, column=5, padx=(0, 6), pady=3)
        ttk.Button(form, text="Export Excel", command=self.export_employees_file).grid(row=0, column=6, padx=(0, 6), pady=3)
        ttk.Button(form, text="Delete Selected", command=self.delete_selected_sales_rep).grid(row=0, column=7, padx=(0, 6), pady=3)
        ttk.Button(form, text="Clear Form", command=self.clear_rep_form).grid(row=0, column=8, padx=(0, 6), pady=3)

        # Search bar for Employees
        rep_search_box = ttk.Frame(self.rep_tab)
        rep_search_box.pack(fill="x", pady=(0, 8))
        ttk.Label(rep_search_box, text="Search:").pack(side="left", padx=(0, 6))
        rep_search_entry = ttk.Entry(rep_search_box, textvariable=self.rep_search_var, width=40)
        rep_search_entry.pack(side="left", padx=(0, 8))
        rep_search_entry.bind("<KeyRelease>", lambda _e: self.refresh_sales_reps_table())
        ttk.Button(rep_search_box, text="Clear", command=lambda: (self.rep_search_var.set(""), self.refresh_sales_reps_table())).pack(side="left")

        columns = ("rep_name", "phone")
        self.rep_tree = ttk.Treeview(self.rep_tab, columns=columns, show="headings", selectmode="browse", height=12)
        self.rep_tree.heading("rep_name", text="Employee Name", command=lambda: self.sort_any_tree(self.rep_tree, "rep_name", False))
        self.rep_tree.heading("phone", text="Phone Number", command=lambda: self.sort_any_tree(self.rep_tree, "phone", False))
        self.rep_tree.column("rep_name", width=320, minwidth=120, anchor="w")
        self.rep_tree.column("phone", width=220, minwidth=120, anchor="w")
        self.rep_tree.pack(fill="both", expand=True)
        self.rep_tree.bind("<<TreeviewSelect>>", self.on_sales_rep_select)

    def import_store_accounts_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Import Store List XLSX",
            filetypes=[
                ("Excel Files", "*.xlsx *.xlsm *.xls"),
                ("All Files", "*.*"),
            ],
        )
        if not path:
            return

        try:
            records: List[Dict[str, str]] = []
            suffix = Path(path).suffix.lower()

            if suffix in {".xlsx", ".xlsm", ".xls"}:
                records = read_xlsx_records(path)
            else:
                messagebox.showerror("Unsupported file", "Please select an Excel file.")
                return

            if not records:
                messagebox.showerror("Empty file", "No store rows found in the selected file.")
                return

            sample = records[0]
            district_col = find_column(sample, ["District", "Distrcit"])
            store_col = find_column(sample, ["Store", "Store Name", "Location"])

            missing = []
            if not district_col:
                missing.append("District")
            if not store_col:
                missing.append("Store")

            if missing:
                messagebox.showerror(
                    "Missing columns",
                    "Missing required column(s): " + ", ".join(missing) +
                    "\n\nRequired headers: District, Store",
                )
                return

            parsed_records: List[Dict[str, str]] = []
            skipped = 0
            for rec in records:
                district = normalize_district(rec.get(district_col, ""))
                store = display_store(rec.get(store_col, ""))

                if not store or not district or district == "Unknown":
                    skipped += 1
                    continue

                parsed_records.append({"District": district, "Store": store})

            imported = self.db.replace_store_accounts(parsed_records)

            self.master_store_records = self.db.store_master_records()
            self.refresh_store_accounts_table()
            self.refresh_entry_form_districts()
            self.populate_status_filters()
            self.populate_audit_filters()

            if self.data_loaded:
                source_name = os.path.basename(self.inventory_path.get().strip()) if self.inventory_path.get().strip() else ""
                self.status_rows, status_summary = build_inventory_status_rows(
                    self.current_inventory_records,
                    self.current_time_sheet_records,
                    master_store_records=self.master_store_records,
                    source_file=source_name,
                )
                self.status_row_by_key = {r.key: r for r in self.status_rows}
                self.db.upsert_inventory_status_rows(self.status_rows)
                self.refresh_status_table()

            messagebox.showinfo(
                "Import complete",
                f"Replaced saved store list with {imported} row(s).\nSkipped {skipped} invalid row(s).\n\nThis list will stay saved until the next upload.",
            )
            self.set_status(f"Store list replaced from file. Saved: {imported}. Skipped: {skipped}.")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Import failed", str(exc))
            self.set_status("Store list import failed.")

    def refresh_store_accounts_table(self) -> None:
        if not hasattr(self, "store_tree"):
            return
        for item in self.store_tree.get_children():
            self.store_tree.delete(item)
        search = self.store_search_var.get().strip().lower() if hasattr(self, "store_search_var") else ""
        for row in self.db.store_accounts():
            district = row.get("District", "")
            store = row.get("Store", "")
            if search and search not in district.lower() and search not in store.lower():
                continue
            iid = row.get("StoreKey", normalize_store(store))
            self.store_tree.insert(
                "",
                "end",
                iid=iid,
                values=(district, store),
            )

    def refresh_entry_form_districts(self) -> None:
        """Refresh all entry-form district comboboxes from the DB."""
        districts = self.db.all_known_districts()
        for combo in (self._store_district_combo, self._wg_district_combo, self._dm_district_combo, self._exclusion_district_combo):
            try:
                combo["values"] = districts
            except Exception:
                pass
        # Also refresh the audit and status filter combos
        if hasattr(self, 'audit_district_combo'):
            self.populate_audit_filters()
        if hasattr(self, 'status_district_combo'):
            self.populate_status_filters()

    def import_employees_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Import Employees XLSX",
            filetypes=[
                ("Excel Files", "*.xlsx *.xlsm *.xls"),
                ("All Files", "*.*"),
            ],
        )
        if not path:
            return

        try:
            records: List[Dict[str, str]] = []
            suffix = Path(path).suffix.lower()

            if suffix in {".xlsx", ".xlsm", ".xls"}:
                records = read_xlsx_records(path)
            else:
                messagebox.showerror("Unsupported file", "Please select an Excel file.")
                return

            if not records:
                messagebox.showerror("Empty file", "No employee rows found in the selected file.")
                return

            sample = records[0]
            name_col = find_column(sample, [
                "Employee Name",
                "Employee",
                "Sales Rep Name",
                "Sales Rep",
                "Rep Name",
                "Salesperson",
                "Name",
                "Full Name",
            ])
            phone_col = find_column(sample, [
                "Phone Number",
                "Phone",
                "Mobile",
                "Mobile Number",
                "Contact",
                "Contact Number",
                "Number",
                "Whatsapp",
                "WhatsApp Number",
            ])

            missing = []
            if not name_col:
                missing.append("Employee Name")
            if not phone_col:
                missing.append("Phone Number")

            if missing:
                messagebox.showerror(
                    "Missing columns",
                    "Missing required column(s): " + ", ".join(missing) +
                    "\n\nRequired headers: Employee Name, Phone Number",
                )
                return

            imported = 0
            skipped = 0
            for rec in records:
                name = safe_text(rec.get(name_col, ""))
                phone = normalize_phone(rec.get(phone_col, ""))

                if not name or not phone:
                    skipped += 1
                    continue

                self.db.save_sales_rep(name, phone)
                imported += 1

            self.refresh_sales_reps_table()
            messagebox.showinfo(
                "Import complete",
                f"Imported or updated {imported} employee phone row(s).\nSkipped {skipped} invalid row(s).\n\nEmployees stay saved until you delete them.",
            )
            self.set_status(f"Imported employees from file. Imported: {imported}. Skipped: {skipped}.")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Import failed", str(exc))
            self.set_status("Employee import failed.")

    def export_employees_file(self) -> None:
        rows = self.db.sales_reps()
        if not rows:
            messagebox.showinfo("No employees", "No employees are saved in the database yet.")
            return

        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"GFH_Employees_{stamp}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Export Employees Excel",
            initialdir=str(EXPORT_DIR),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not path:
            return

        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Name", "Phone Number"])
        for r_idx, row in enumerate(rows, start=2):
            ws.cell(row=r_idx, column=1, value=row.get("Rep Name", ""))
            ws.cell(row=r_idx, column=2, value=row.get("Phone", ""))
        wb.save(output_path)

        self.set_status(f"Exported employee list: {output_path}")
        messagebox.showinfo("Export complete", f"Employee list exported:\n{output_path}")

    def refresh_sales_reps_table(self) -> None:
        if not hasattr(self, "rep_tree"):
            return
        for item in self.rep_tree.get_children():
            self.rep_tree.delete(item)
        search = self.rep_search_var.get().strip().lower() if hasattr(self, "rep_search_var") else ""
        for row in self.db.sales_reps():
            rep_name = row.get("Rep Name", "")
            phone = row.get("Phone", "")
            if search and search not in rep_name.lower() and search not in phone.lower():
                continue
            iid = row.get("RepKey", person_name_key(rep_name))
            self.rep_tree.insert("", "end", iid=iid, values=(rep_name, phone))

    def save_store_account_from_form(self) -> None:
        try:
            self.db.save_store_account(
                self.store_district_var.get(),
                self.store_name_var.get(),
                "",
                "",
                "",
            )
            self.master_store_records = self.db.store_master_records()
            self.refresh_store_accounts_table()
            self.refresh_entry_form_districts()
            self.populate_status_filters()
            self.populate_audit_filters()
            if self.data_loaded:
                source_name = os.path.basename(self.inventory_path.get().strip()) if self.inventory_path.get().strip() else ""
                self.status_rows, status_summary = build_inventory_status_rows(
                    self.current_inventory_records,
                    self.current_time_sheet_records,
                    master_store_records=self.master_store_records,
                    source_file=source_name,
                )
                self.status_row_by_key = {r.key: r for r in self.status_rows}
                self.db.upsert_inventory_status_rows(self.status_rows)
                self.refresh_status_table()
            self.set_status(f"Saved store: {self.store_district_var.get()} | {self.store_name_var.get()}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def delete_selected_store_account(self) -> None:
        selected = self.store_tree.selection() if hasattr(self, "store_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select a store first.")
            return
        if not messagebox.askyesno("Delete store", "Delete the selected store?"):
            return
        self.db.delete_store_account(selected[0])
        self.master_store_records = self.db.store_master_records()
        self.refresh_store_accounts_table()
        self.refresh_entry_form_districts()
        self.clear_store_form()
        self.set_status("Deleted selected store.")

    def clear_store_form(self) -> None:
        self.store_district_var.set("")
        self.store_name_var.set("")

    def on_store_account_select(self, _event=None) -> None:
        selected = self.store_tree.selection() if hasattr(self, "store_tree") else []
        if not selected:
            return
        store_key = selected[0]
        for row in self.db.store_accounts():
            if row.get("StoreKey") == store_key:
                self.store_district_var.set(row.get("District", ""))
                self.store_name_var.set(row.get("Store", ""))
                return

    def save_sales_rep_from_form(self) -> None:
        try:
            self.db.save_sales_rep(self.rep_name_var.get(), self.rep_phone_var.get())
            self.refresh_sales_reps_table()
            self.set_status(f"Saved employee phone: {self.rep_name_var.get()} | {self.rep_phone_var.get()}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def delete_selected_sales_rep(self) -> None:
        selected = self.rep_tree.selection() if hasattr(self, "rep_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select an employee first.")
            return
        if not messagebox.askyesno("Delete employee", "Delete the selected employee phone record from the database?"):
            return
        self.db.delete_sales_rep(selected[0])
        self.refresh_sales_reps_table()
        self.clear_rep_form()
        self.set_status("Deleted selected employee from database.")

    def clear_rep_form(self) -> None:
        self.rep_name_var.set("")
        self.rep_phone_var.set("")

    def on_sales_rep_select(self, _event=None) -> None:
        selected = self.rep_tree.selection() if hasattr(self, "rep_tree") else []
        if not selected:
            return
        rep_key = selected[0]
        for row in self.db.sales_reps():
            if row.get("RepKey") == rep_key:
                self.rep_name_var.set(row.get("Rep Name", ""))
                self.rep_phone_var.set(row.get("Phone", ""))
                return

    def _build_dm_tab(self) -> None:
        # Side-by-side layout using PanedWindow
        pw = ttk.PanedWindow(self.dm_tab, orient="horizontal")
        pw.pack(fill="both", expand=True)

        # ── Left: WhatsApp Group Names ──────────────────────────────────────
        left = ttk.Frame(pw, padding=4)
        pw.add(left, weight=1)

        wg_form = ttk.LabelFrame(left, text="WhatsApp Group Names", padding=8)
        wg_form.pack(fill="x", pady=(0, 6))

        wg_form_row = ttk.Frame(wg_form)
        wg_form_row.pack(fill="x")
        ttk.Label(wg_form_row, text="District:").pack(side="left", padx=(0, 4))
        self._wg_district_combo = ttk.Combobox(
            wg_form_row,
            textvariable=self.wg_district_var,
            values=self.db.all_known_districts(),
            width=18,
        )
        self._wg_district_combo.pack(side="left", padx=(0, 8))
        ttk.Label(wg_form_row, text="Group Name:").pack(side="left", padx=(0, 4))
        ttk.Entry(wg_form_row, textvariable=self.wg_group_name_var, width=28).pack(side="left", padx=(0, 8))

        wg_btn_row = ttk.Frame(wg_form)
        wg_btn_row.pack(fill="x", pady=(4, 0))
        ttk.Button(wg_btn_row, text="Save", command=self.save_whatsapp_group_from_form).pack(side="left", padx=(0, 4))
        ttk.Button(wg_btn_row, text="Delete", command=self.delete_selected_whatsapp_group).pack(side="left", padx=(0, 4))
        ttk.Button(wg_btn_row, text="Clear", command=self.clear_wg_form).pack(side="left", padx=(0, 4))

        wg_columns = ("district", "group_name")
        self.wg_tree = ttk.Treeview(left, columns=wg_columns, show="headings", selectmode="browse", height=12)
        self.wg_tree.heading("district", text="District", command=lambda: self.sort_any_tree(self.wg_tree, "district", False))
        self.wg_tree.heading("group_name", text="WhatsApp Group Name", command=lambda: self.sort_any_tree(self.wg_tree, "group_name", False))
        self.wg_tree.column("district", width=160, minwidth=100, anchor="w")
        self.wg_tree.column("group_name", width=280, minwidth=160, anchor="w")
        self.wg_tree.pack(fill="both", expand=True)
        self.wg_tree.bind("<<TreeviewSelect>>", self.on_whatsapp_group_select)

        # ── Right: District Manager WhatsApp Tags ───────────────────────────
        right = ttk.Frame(pw, padding=4)
        pw.add(right, weight=1)

        form = ttk.LabelFrame(right, text="DM WhatsApp Tags", padding=8)
        form.pack(fill="x", pady=(0, 6))

        dm_form_row = ttk.Frame(form)
        dm_form_row.pack(fill="x")
        ttk.Label(dm_form_row, text="District:").pack(side="left", padx=(0, 4))
        self._dm_district_combo = ttk.Combobox(
            dm_form_row,
            textvariable=self.dm_district_var,
            values=self.db.all_known_districts(),
            width=18,
        )
        self._dm_district_combo.pack(side="left", padx=(0, 8))
        ttk.Label(dm_form_row, text="DM Name:").pack(side="left", padx=(0, 4))
        ttk.Entry(dm_form_row, textvariable=self.dm_name_var, width=22).pack(side="left", padx=(0, 8))
        ttk.Label(dm_form_row, text="Phone:").pack(side="left", padx=(0, 4))
        ttk.Entry(dm_form_row, textvariable=self.dm_phone_var, width=18).pack(side="left", padx=(0, 8))

        dm_btn_row = ttk.Frame(form)
        dm_btn_row.pack(fill="x", pady=(4, 0))
        ttk.Button(dm_btn_row, text="Save", command=self.save_district_manager_from_form).pack(side="left", padx=(0, 4))
        ttk.Button(dm_btn_row, text="Delete", command=self.delete_selected_district_manager).pack(side="left", padx=(0, 4))
        ttk.Button(dm_btn_row, text="Clear", command=self.clear_dm_form).pack(side="left", padx=(0, 4))

        columns = ("district", "dm_name", "phone")
        self.dm_tree = ttk.Treeview(right, columns=columns, show="headings", selectmode="browse", height=12)
        self.dm_tree.heading("district", text="District", command=lambda: self.sort_any_tree(self.dm_tree, "district", False))
        self.dm_tree.heading("dm_name", text="DM Name", command=lambda: self.sort_any_tree(self.dm_tree, "dm_name", False))
        self.dm_tree.heading("phone", text="Phone Number", command=lambda: self.sort_any_tree(self.dm_tree, "phone", False))
        self.dm_tree.column("district", width=160, minwidth=100, anchor="w")
        self.dm_tree.column("dm_name", width=220, minwidth=120, anchor="w")
        self.dm_tree.column("phone", width=180, minwidth=120, anchor="w")
        self.dm_tree.pack(fill="both", expand=True)
        self.dm_tree.bind("<<TreeviewSelect>>", self.on_district_manager_select)

    def refresh_district_managers_table(self) -> None:
        if not hasattr(self, "dm_tree"):
            return
        for item in self.dm_tree.get_children():
            self.dm_tree.delete(item)
        for row in self.db.district_managers():
            iid = normalize_district(row.get("District", ""))
            self.dm_tree.insert("", "end", iid=iid, values=(row.get("District", ""), row.get("DM Name", ""), row.get("Phone", "")))

    def save_district_manager_from_form(self) -> None:
        try:
            self.db.save_district_manager(self.dm_district_var.get(), self.dm_name_var.get(), self.dm_phone_var.get())
            self.refresh_district_managers_table()
            self.refresh_entry_form_districts()
            self.set_status(f"Saved district DM: {self.dm_district_var.get()} | {self.dm_phone_var.get()}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def delete_selected_district_manager(self) -> None:
        selected = self.dm_tree.selection() if hasattr(self, "dm_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select a district DM first.")
            return
        if not messagebox.askyesno("Delete district DM", "Delete the selected district DM phone record?"):
            return
        self.db.delete_district_manager(selected[0])
        self.refresh_district_managers_table()
        self.refresh_entry_form_districts()
        self.clear_dm_form()
        self.set_status("Deleted selected district DM.")

    def clear_dm_form(self) -> None:
        self.dm_district_var.set("")
        self.dm_name_var.set("")
        self.dm_phone_var.set("")

    def refresh_whatsapp_groups_table(self) -> None:
        if not hasattr(self, "wg_tree"):
            return
        for item in self.wg_tree.get_children():
            self.wg_tree.delete(item)
        for row in self.db.whatsapp_groups():
            iid = normalize_district(row.get("District", ""))
            self.wg_tree.insert("", "end", iid=iid, values=(row.get("District", ""), row.get("Group Name", "")))

    def save_whatsapp_group_from_form(self) -> None:
        try:
            self.db.save_whatsapp_group(self.wg_district_var.get(), self.wg_group_name_var.get())
            self.refresh_whatsapp_groups_table()
            self.refresh_entry_form_districts()
            self.set_status(f"Saved WhatsApp group: {self.wg_district_var.get()} | {self.wg_group_name_var.get()}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def delete_selected_whatsapp_group(self) -> None:
        selected = self.wg_tree.selection() if hasattr(self, "wg_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select a WhatsApp group first.")
            return
        if not messagebox.askyesno("Delete WhatsApp group", "Delete the selected WhatsApp group record?"):
            return
        self.db.delete_whatsapp_group(selected[0])
        self.refresh_whatsapp_groups_table()
        self.refresh_entry_form_districts()
        self.clear_wg_form()
        self.set_status("Deleted selected WhatsApp group.")

    def clear_wg_form(self) -> None:
        self.wg_district_var.set("")
        self.wg_group_name_var.set("")

    def on_whatsapp_group_select(self, _event=None) -> None:
        selected = self.wg_tree.selection() if hasattr(self, "wg_tree") else []
        if not selected:
            return
        district = selected[0]
        for row in self.db.whatsapp_groups():
            if normalize_district(row.get("District", "")) == normalize_district(district):
                self.wg_district_var.set(row.get("District", ""))
                self.wg_group_name_var.set(row.get("Group Name", ""))
                return

    def on_district_manager_select(self, _event=None) -> None:
        selected = self.dm_tree.selection() if hasattr(self, "dm_tree") else []
        if not selected:
            return
        district = selected[0]
        for row in self.db.district_managers():
            if normalize_district(row.get("District", "")) == normalize_district(district):
                self.dm_district_var.set(row.get("District", ""))
                self.dm_name_var.set(row.get("DM Name", ""))
                self.dm_phone_var.set(row.get("Phone", ""))
                return

    def _build_exclusion_tab(self) -> None:
        form = ttk.LabelFrame(self.exclusion_tab, text="Exclude Devices From Variance Images", padding=10)
        form.pack(fill="x", pady=(0, 8))
        ttk.Label(form, text="District").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=3)
        self._exclusion_district_combo = ttk.Combobox(
            form,
            textvariable=self.exclusion_district_var,
            values=self.db.all_known_districts(),
            width=16,
        )
        self._exclusion_district_combo.grid(row=0, column=1, sticky="w", padx=(0, 8), pady=3)
        ttk.Label(form, text="Product").grid(row=0, column=2, sticky="w", padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=self.exclusion_product_var, width=30).grid(row=0, column=3, sticky="w", padx=(0, 12), pady=3)
        ttk.Label(form, text="IMEI").grid(row=0, column=4, sticky="w", padx=(0, 6), pady=3)
        imei_entry = ttk.Entry(form, textvariable=self.exclusion_imei_var, width=22)
        imei_entry.grid(row=0, column=5, sticky="w", padx=(0, 12), pady=3)
        imei_entry.bind("<FocusOut>", self.lookup_product_for_exclusion_imei)
        imei_entry.bind("<Return>", self.lookup_product_for_exclusion_imei)
        ttk.Label(form, text="Comments").grid(row=0, column=6, sticky="w", padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=self.exclusion_comments_var, width=36).grid(row=0, column=7, sticky="w", padx=(0, 8), pady=3)
        ttk.Button(form, text="Find Product", command=self.lookup_product_for_exclusion_imei).grid(row=0, column=8, padx=(0, 4), pady=3)
        ttk.Button(form, text="Save", command=self.save_device_exclusion_from_form).grid(row=0, column=9, padx=(0, 4), pady=3)
        ttk.Button(form, text="Update Comment", command=self.update_selected_exclusion_comment).grid(row=0, column=10, padx=(0, 4), pady=3)
        ttk.Button(form, text="Import XLSX", command=self.import_excluded_imeis_file).grid(row=0, column=11, padx=(0, 4), pady=3)
        ttk.Button(form, text="Export Excel", command=self.export_excluded_imeis_file).grid(row=0, column=12, padx=(0, 4), pady=3)
        ttk.Button(form, text="Delete", command=self.delete_selected_device_exclusion).grid(row=0, column=13, padx=(0, 4), pady=3)
        ttk.Button(form, text="Delete All", command=self.delete_all_device_exclusions_from_ui).grid(row=0, column=14, padx=(0, 4), pady=3)
        ttk.Button(form, text="Clear", command=self.clear_device_exclusion_form).grid(row=0, column=15, padx=(0, 4), pady=3)

        # Search bar for Excluded Devices
        exclusion_search_box = ttk.Frame(self.exclusion_tab)
        exclusion_search_box.pack(fill="x", pady=(0, 8))
        ttk.Label(exclusion_search_box, text="Search:").pack(side="left", padx=(0, 6))
        exclusion_search_entry = ttk.Entry(exclusion_search_box, textvariable=self.exclusion_search_var, width=40)
        exclusion_search_entry.pack(side="left", padx=(0, 8))
        exclusion_search_entry.bind("<KeyRelease>", lambda _e: self.refresh_device_exclusions_table())
        ttk.Button(exclusion_search_box, text="Clear", command=lambda: (self.exclusion_search_var.set(""), self.refresh_device_exclusions_table())).pack(side="left")

        columns = ("district", "product", "imei", "comments")
        self.exclusion_tree = ttk.Treeview(self.exclusion_tab, columns=columns, show="headings", selectmode="browse", height=12)
        self.exclusion_tree.heading("district", text="District", command=lambda: self.sort_any_tree(self.exclusion_tree, "district", False))
        self.exclusion_tree.heading("product", text="Product", command=lambda: self.sort_any_tree(self.exclusion_tree, "product", False))
        self.exclusion_tree.heading("imei", text="IMEI", command=lambda: self.sort_any_tree(self.exclusion_tree, "imei", False))
        self.exclusion_tree.heading("comments", text="Comments", command=lambda: self.sort_any_tree(self.exclusion_tree, "comments", False))
        self.exclusion_tree.column("district", width=160, minwidth=120, anchor="w")
        self.exclusion_tree.column("product", width=400, minwidth=160, anchor="w")
        self.exclusion_tree.column("imei", width=220, minwidth=120, anchor="w")
        self.exclusion_tree.column("comments", width=420, minwidth=160, anchor="w")
        self.exclusion_tree.pack(fill="both", expand=True)
        self.exclusion_tree.bind("<<TreeviewSelect>>", self.on_device_exclusion_select)
        self.exclusion_tree.bind("<Double-1>", self.edit_selected_exclusion_comment)

    def delete_all_device_exclusions_from_ui(self) -> None:
        if not messagebox.askyesno("Delete All Exclusions", "Are you sure you want to completely remove ALL device exclusions from the database? This cannot be undone."):
            return
        self.db.delete_all_device_exclusions()
        self.refresh_device_exclusions_table()
        self.refresh_entry_form_districts()
        self.refresh_table()
        self.clear_device_exclusion_form()
        self.set_status("Deleted all device exclusions.")

    def refresh_device_exclusions_table(self) -> None:
        if not hasattr(self, "exclusion_tree"):
            return
        for item in self.exclusion_tree.get_children():
            self.exclusion_tree.delete(item)
        search = self.exclusion_search_var.get().strip().lower() if hasattr(self, "exclusion_search_var") else ""
        for row in self.db.device_exclusions():
            district = row.get("District", "")
            product = row.get("Product", "")
            imei = row.get("IMEI", "")
            comments = row.get("Comments", "")
            if search and search not in district.lower() and search not in product.lower() and search not in imei.lower() and search not in comments.lower():
                continue
            iid = row.get("RuleKey", device_rule_key(district + "|" + product + "|" + imei))
            self.exclusion_tree.insert(
                "",
                "end",
                iid=iid,
                values=(district, product, imei, comments),
            )

    def product_for_imei_from_loaded_rows(self, imei: str) -> str:
        imei_key = device_rule_key(imei)
        if not imei_key:
            return ""

        # Prefer current loaded variance rows.
        for row in self.db.get_rows_by_keys(self.loaded_keys) if self.loaded_keys else []:
            if device_rule_key(row.imei) == imei_key:
                return safe_text(row.product)

        # Fallback to currently read inventory records if available.
        for rec in self.current_inventory_records:
            for key, value in rec.items():
                if device_rule_key(value) == imei_key:
                    product_col = find_column(rec, ["Product", "Product Name", "Device", "Model", "Item", "SKU Description"])
                    if product_col:
                        return safe_text(rec.get(product_col, ""))
        return ""

    def lookup_product_for_exclusion_imei(self, _event=None):
        imei = self.exclusion_imei_var.get()
        if not safe_text(imei):
            return "break"
        product = self.product_for_imei_from_loaded_rows(imei)
        if product:
            self.exclusion_product_var.set(product)
            self.set_status(f"Product found for IMEI {imei}: {product}")
        else:
            self.set_status(f"No product found for IMEI {imei}. Product field left unchanged.")
        return "break"

    def save_device_exclusion_from_form(self) -> None:
        try:
            district = self.exclusion_district_var.get()
            imei = self.exclusion_imei_var.get()
            product = self.exclusion_product_var.get()
            comments = self.exclusion_comments_var.get()

            if safe_text(imei) and not safe_text(product):
                found_product = self.product_for_imei_from_loaded_rows(imei)
                if found_product:
                    product = found_product
                    self.exclusion_product_var.set(found_product)

            # IMEI exclusions stay global and ignore district.
            if safe_text(imei):
                district = ""
                self.exclusion_district_var.set("")

            selected_key = self.selected_exclusion_key_var.get().strip()
            if selected_key:
                self.db.delete_device_exclusion(selected_key)

            self.db.save_device_exclusion(product, imei, comments, district)
            self.selected_exclusion_key_var.set("")
            self.refresh_device_exclusions_table()
            self.refresh_entry_form_districts()
            self.refresh_table()
            self.set_status(
                f"Saved exclusion. District: {safe_text(district) or '-'} | Product: {safe_text(product) or '-'} | IMEI: {safe_text(imei) or '-'} | Comments: {safe_text(comments) or '-'}"
            )
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def update_selected_exclusion_comment(self) -> None:
        selected = self.exclusion_tree.selection() if hasattr(self, "exclusion_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select an exclusion row first.")
            return
        rule_key = selected[0]
        try:
            self.db.update_device_exclusion_comment(rule_key, self.exclusion_comments_var.get())
            self.selected_exclusion_key_var.set(rule_key)
            self.refresh_device_exclusions_table()
            self.refresh_table()
            self.set_status("Updated selected exclusion comment.")
        except Exception as exc:
            messagebox.showerror("Update failed", str(exc))

    def edit_selected_exclusion_comment(self, _event=None):
        selected = self.exclusion_tree.selection() if hasattr(self, "exclusion_tree") else []
        if not selected:
            return "break"

        current_comment = self.exclusion_comments_var.get()
        new_comment = simpledialog.askstring(
            "Edit Comments",
            "Enter comments for selected exclusion.",
            initialvalue=current_comment,
            parent=self,
        )
        if new_comment is None:
            return "break"

        self.exclusion_comments_var.set(new_comment)
        self.update_selected_exclusion_comment()
        return "break"

    def import_excluded_imeis_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Import Excluded Devices XLSX",
            filetypes=[
                ("Excel Files", "*.xlsx *.xlsm *.xls"),
                ("All Files", "*.*"),
            ],
        )
        if not path:
            return

        try:
            records: List[Dict[str, str]] = []
            suffix = Path(path).suffix.lower()
            if suffix in {".xlsx", ".xlsm", ".xls"}:
                records = read_xlsx_records(path)
            else:
                messagebox.showerror("Unsupported file", "Please select an Excel file.")
                return

            if not records:
                messagebox.showerror("Empty file", "No exclusion rows found in the selected file.")
                return

            sample = records[0]
            district_col = find_column(sample, ["District", "Market", "Area"])
            product_col = find_column(sample, ["Product", "Product Name", "Device", "Model", "Item"])
            imei_col = find_column(sample, ["IMEI", "IMEI Number", "ESN", "ESN Number", "Serial", "Serial 1", "Device ID"])
            comments_col = find_column(sample, ["Comments", "Comment", "Notes", "Note", "Reason"])

            if not product_col and not imei_col:
                messagebox.showerror(
                    "Missing columns",
                    "Import file needs at least Product or IMEI. Supported headers: District, Product, IMEI, Comments.",
                )
                return

            imported = 0
            skipped = 0
            for rec in records:
                district = normalize_district(rec.get(district_col, "")) if district_col else ""
                product = safe_text(rec.get(product_col, "")) if product_col else ""
                imei = safe_text(rec.get(imei_col, "")) if imei_col else ""
                comments = safe_text(rec.get(comments_col, "")) if comments_col else ""

                if imei and not product:
                    product = self.product_for_imei_from_loaded_rows(imei)

                # IMEI exclusions ignore district. Product exclusions require district.
                if imei:
                    district = ""
                elif product and (not district or district == "Unknown"):
                    skipped += 1
                    continue

                if not product and not imei:
                    skipped += 1
                    continue

                self.db.save_device_exclusion(product, imei, comments, district)
                imported += 1

            self.refresh_device_exclusions_table()
            self.refresh_table()
            messagebox.showinfo("Import complete", f"Imported {imported} exclusion row(s).\nSkipped {skipped} invalid row(s).\n\nProduct exclusions require District. IMEI exclusions are global.")
            self.set_status(f"Imported excluded devices. Imported: {imported}. Skipped: {skipped}.")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Import failed", str(exc))
            self.set_status("Excluded devices import failed.")

    def export_excluded_imeis_file(self) -> None:
        rows = self.db.device_exclusions()
        if not rows:
            messagebox.showinfo("No exclusions", "No excluded devices are saved.")
            return
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"GFH_Excluded_Devices_{stamp}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Export Excluded Devices",
            initialdir=str(EXPORT_DIR),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not path:
            return
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["District", "Product", "IMEI", "Comments"])
        for r_idx, row in enumerate(rows, start=2):
            ws.cell(row=r_idx, column=1, value=row.get("District", ""))
            ws.cell(row=r_idx, column=2, value=row.get("Product", ""))
            
            imei_cell = ws.cell(row=r_idx, column=3, value=row.get("IMEI", ""))
            imei_cell.data_type = 's'
            imei_cell.number_format = '@'
            
            ws.cell(row=r_idx, column=4, value=row.get("Comments", ""))
        wb.save(output_path)

        self.set_status(f"Exported excluded devices: {output_path}")
        messagebox.showinfo("Export complete", f"Excluded devices exported:\n{output_path}")

    def delete_selected_device_exclusion(self) -> None:
        selected = self.exclusion_tree.selection() if hasattr(self, "exclusion_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select a device exclusion first.")
            return
        if not messagebox.askyesno("Delete exclusion", "Delete the selected device exclusion from the database?"):
            return
        self.db.delete_device_exclusion(selected[0])
        self.refresh_device_exclusions_table()
        self.refresh_entry_form_districts()
        self.refresh_table()
        self.clear_device_exclusion_form()
        self.set_status("Deleted selected device exclusion.")

    def clear_device_exclusion_form(self) -> None:
        self.exclusion_district_var.set("")
        self.exclusion_product_var.set("")
        self.exclusion_imei_var.set("")
        self.exclusion_comments_var.set("")
        self.selected_exclusion_key_var.set("")

    def on_device_exclusion_select(self, _event=None) -> None:
        selected = self.exclusion_tree.selection() if hasattr(self, "exclusion_tree") else []
        if not selected:
            return
        rule_key = selected[0]
        self.selected_exclusion_key_var.set(rule_key)
        for row in self.db.device_exclusions():
            if row.get("RuleKey") == rule_key:
                self.exclusion_district_var.set(row.get("District", ""))
                self.exclusion_product_var.set(row.get("Product", ""))
                self.exclusion_imei_var.set(row.get("IMEI", ""))
                self.exclusion_comments_var.set(row.get("Comments", ""))
                return

    def filter_excluded_variance_rows(self, rows: List[VarianceRow]) -> List[VarianceRow]:
        if not rows:
            return []
        return [row for row in rows if not self.db.is_device_excluded(row.district, row.product, row.imei)]

    def _file_row(self, parent, row: int, label: str, var: tk.StringVar, command) -> None:
        ttk.Label(parent, text=label, width=30).grid(row=row, column=0, sticky="w", pady=3)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=3)
        ttk.Button(parent, text="Browse", command=command).grid(row=row, column=2, padx=(8, 0), pady=3)

    def pick_inventory(self) -> None:
        path = filedialog.askopenfilename(title="Select Inventory_Count_Result_Details", filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")])
        if path:
            self.inventory_path.set(path)

    def pick_time_sheet(self) -> None:
        path = filedialog.askopenfilename(title="Select Employee_Time_Sheet", filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")])
        if path:
            self.time_sheet_path.set(path)

    def set_status(self, text: str) -> None:
        self.status_text.set(text)
        self.update_idletasks()

    def load_variances(self) -> None:
        inventory = self.inventory_path.get().strip()
        time_sheet = self.time_sheet_path.get().strip()
        if not inventory or not os.path.exists(inventory):
            messagebox.showerror("Missing file", "Select Inventory_Count_Result_Details.xlsx first.")
            return
        if not time_sheet or not os.path.exists(time_sheet):
            messagebox.showerror("Missing file", "Select Employee_Time_Sheet.xlsx first.")
            return
        try:
            self.clear_current_ui(silent=True)
            self.set_status("Reading Excel files...")
            inv_records = read_xlsx_records(inventory)
            ts_records = read_xlsx_records(time_sheet)
            self.current_inventory_records = inv_records
            self.current_time_sheet_records = ts_records
            self.master_store_records = self.db.store_master_records()
            self.set_status("Building Inventory Status and Variance Audit rows...")

            self.status_rows, status_summary = build_inventory_status_rows(inv_records, ts_records, master_store_records=self.master_store_records, source_file=os.path.basename(inventory))
            self.status_row_by_key = {r.key: r for r in self.status_rows}
            self.db.upsert_inventory_status_rows(self.status_rows)
            rows, variance_summary = extract_variances(inv_records, ts_records, master_store_records=self.master_store_records, source_file=os.path.basename(inventory))
            self.db.upsert_rows(rows)
            self.loaded_keys = {row.key for row in rows}
            self.data_loaded = True

            self.populate_status_filters()
            self.populate_audit_filters()
            self.refresh_status_table()
            self.refresh_table()

            self.summary_text.set(
                f"Inventory Status. Completed: {status_summary['completed']}   Pending: {status_summary['pending']}   "
                f"Variance Audit: {len(rows)}   SIMs skipped: {variance_summary['skipped_sims']}"
            )
            self.set_status(
                f"Loaded current session. Inventory Status rows: {len(self.status_rows)}. Variances: {len(rows)}. "
                f"Inventory rows read: {variance_summary['raw_inventory_rows']}. Latest rows used: {variance_summary['latest_inventory_rows']}. "
                f"Stale rows skipped: {variance_summary['stale_inventory_rows']}. SIM rows skipped: {variance_summary['skipped_sims']}."
            )
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Load failed", str(exc))
            self.set_status("Load failed. Check the Excel columns and try again.")

    def add_store_prompt(self) -> None:
        district = simpledialog.askstring(
            "Add Store",
            "Enter district name.",
            parent=self,
        )
        if district is None:
            return
        district = normalize_district(district)
        if district == "Unknown":
            messagebox.showerror("Invalid district", "Please enter a valid district name.")
            return

        store = simpledialog.askstring("Add Store", "Enter store name.", parent=self)
        if store is None:
            return
        store = display_store(store)
        if not store:
            messagebox.showerror("Invalid store", "Store name is required.")
            return

        self.db.save_store_account(district, store, "", "", "")
        self.master_store_records = self.db.store_master_records()
        self.refresh_store_accounts_table()
        self.refresh_entry_form_districts()

        if self.data_loaded:
            source_name = os.path.basename(self.inventory_path.get().strip()) if self.inventory_path.get().strip() else ""
            self.status_rows, status_summary = build_inventory_status_rows(
                self.current_inventory_records,
                self.current_time_sheet_records,
                master_store_records=self.master_store_records,
                source_file=source_name,
            )
            self.status_row_by_key = {r.key: r for r in self.status_rows}
            self.populate_status_filters()
            self.refresh_status_table()
            self.summary_text.set(
                f"Inventory Audit Status. Completed: {status_summary['completed']}   Pending: {status_summary['pending']}   Variance Audit: {len(self.loaded_keys)}"
            )
        else:
            self.populate_status_filters()

        messagebox.showinfo("Store saved", f"Saved store.\nDistrict: {district}\nStore: {store}")
        self.set_status(f"Added store: {district} | {store}")

    def populate_status_filters(self) -> None:
        districts = set()
        for r in self.status_rows:
            d = normalize_district(r.district or "Unknown")
            if d and d != "Unknown":
                districts.add(d)
        for d in self.db.all_known_districts():
            if d and d != "Unknown":
                districts.add(d)
        district_values = ["All Districts"] + sorted(districts)
        self.status_district_combo["values"] = district_values
        if self.status_district_filter.get() not in district_values:
            self.status_district_filter.set("All Districts")
        self.update_status_store_filter_values()

    def update_status_store_filter_values(self) -> None:
        selected_district = self.status_district_filter.get().strip()
        rows = self.status_rows
        if selected_district and selected_district != "All Districts":
            rows = [r for r in rows if normalize_district(r.district) == normalize_district(selected_district)]
        stores = sorted({r.store for r in rows if r.store})
        values = ["All Stores"] + stores
        self.status_store_combo["values"] = values
        if self.status_store_filter.get() not in values:
            self.status_store_filter.set("All Stores")

    def populate_audit_filters(self) -> None:
        districts = set()
        rows = self.db.get_rows_by_keys(self.loaded_keys) if self.loaded_keys else []
        for r in rows:
            d = normalize_district(r.district or "Unknown")
            if d and d != "Unknown":
                districts.add(d)
        for d in self.db.all_known_districts():
            if d and d != "Unknown":
                districts.add(d)
        district_values = ["All Districts"] + sorted(districts)
        self.audit_district_combo["values"] = district_values
        if self.audit_district_filter.get() not in district_values:
            self.audit_district_filter.set("All Districts")
        # Also populate the final district combo
        if hasattr(self, "final_district_combo"):
            self.final_district_combo["values"] = district_values
            if self.final_district_var.get() not in district_values:
                self.final_district_var.set("All Districts")
        self.update_audit_store_filter_values()

    def update_audit_store_filter_values(self) -> None:
        rows = self.db.get_rows_by_keys(self.loaded_keys) if self.loaded_keys else []
        district = self.audit_district_filter.get().strip()
        if district and district != "All Districts":
            rows = [r for r in rows if normalize_district(r.district) == normalize_district(district)]
        stores = sorted({r.store for r in rows if r.store})
        values = ["All Stores"] + stores
        self.audit_store_combo["values"] = values
        if self.audit_store_filter.get() not in values:
            self.audit_store_filter.set("All Stores")

    def on_status_district_change(self, _event=None) -> None:
        self.update_status_store_filter_values()
        self.auto_check_status_rows()
        self.refresh_status_table()

    def on_status_store_change(self, _event=None) -> None:
        self.auto_check_status_rows()
        self.refresh_status_table()

    def on_audit_district_change(self, _event=None) -> None:
        self.update_audit_store_filter_values()
        self.auto_check_audit_rows()
        self.refresh_table()

    def on_audit_store_change(self, _event=None) -> None:
        self.auto_check_audit_rows()
        self.refresh_table()

    def matching_status_rows(self) -> List[InventoryStatusRow]:
        rows = list(self.status_rows)
        district = self.status_district_filter.get().strip()
        store = self.status_store_filter.get().strip()
        if district and district != "All Districts":
            rows = [r for r in rows if normalize_district(r.district) == normalize_district(district)]
        if store and store != "All Stores":
            rows = [r for r in rows if r.store == store]

        any_q = safe_text(self.status_search_any_var.get()).lower()
        district_q = safe_text(self.status_search_district_var.get()).lower()
        store_q = safe_text(self.status_search_store_var.get()).lower()
        status_q = safe_text(self.status_search_status_var.get()).lower()
        rep_q = safe_text(self.status_search_rep_var.get()).lower()

        if district_q:
            rows = [r for r in rows if district_q in safe_text(r.district).lower()]
        if store_q:
            rows = [r for r in rows if store_q in safe_text(r.store).lower()]
        if status_q:
            rows = [r for r in rows if status_q in safe_text(r.status).lower()]
        if rep_q:
            rows = [r for r in rows if rep_q in safe_text(r.rep_name).lower()]
        if any_q:
            rows = [
                r for r in rows
                if any_q in " ".join([
                    safe_text(r.district),
                    safe_text(r.store),
                    safe_text(r.status),
                    safe_text(r.rep_name),
                ]).lower()
            ]
        return rows

    def clear_status_search(self) -> None:
        self.status_search_any_var.set("")
        self.status_search_district_var.set("")
        self.status_search_store_var.set("")
        self.status_search_status_var.set("")
        self.status_search_rep_var.set("")
        self.refresh_status_table()

    def status_row_tag(self, row: InventoryStatusRow, state_map: Optional[Dict[str, Dict[str, str]]] = None) -> str:
        state_map = state_map if state_map is not None else self.db.inventory_status_state_map()
        state = state_map.get(normalize_store(row.store), {})
        status_norm = normalize_header(row.status)
        previous_norm = normalize_header(state.get("Previous Status", ""))
        sent_at = safe_text(state.get("Last Sent At", ""))

        if status_norm != "completed":
            return "status_pending"
        if sent_at:
            return "status_completed_sent"
        if previous_norm == "pending":
            return "status_completed_after_update"
        return "status_completed"

    def auto_check_pending_status_rows(self) -> None:
        rows = [r for r in self.matching_status_rows() if normalize_header(r.status) != "completed"]
        self.status_checked_keys = {r.key for r in rows}
        self.refresh_status_table()
        self.set_status(f"Prepared {len(rows)} pending inventory status row(s) for sending.")

    def auto_check_status_rows(self) -> None:
        rows = self.matching_status_rows()
        self.status_checked_keys = {r.key for r in rows}
        self.refresh_status_table()
        self.set_status(f"Prepared {len(rows)} inventory status row(s) for sending. Uncheck any row you want to skip.")

    def clear_status_checkmarks(self) -> None:
        self.status_checked_keys.clear()
        self.refresh_status_table()
        self.set_status("Inventory status checkmarks cleared.")

    def refresh_status_table(self) -> None:
        for item in self.status_tree.get_children():
            self.status_tree.delete(item)
        self.status_key_by_iid.clear()
        if not self.data_loaded:
            return
        rows = self.matching_status_rows()
        rows.sort(key=lambda r: (r.district, r.store, r.rep_name, r.status))
        state_map = self.db.inventory_status_state_map()
        for row in rows:
            checkbox = "☑ Send" if row.key in self.status_checked_keys else "☐ Skip"
            tag = self.status_row_tag(row, state_map)
            iid = self.status_tree.insert("", "end", values=(row.district, row.store, row.status, row.rep_name, checkbox), tags=(tag,))
            self.status_key_by_iid[iid] = row.key
        checked = len([r for r in rows if r.key in self.status_checked_keys])
        self.set_status(f"Inventory Status rows shown: {len(rows)}. Checked for send: {checked}.")

    def on_status_tree_click(self, event) -> None:
        region = self.status_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col_id = self.status_tree.identify_column(event.x)
        columns = self.status_tree["columns"]
        try:
            col_name = columns[int(col_id.replace("#", "")) - 1]
        except Exception:
            return
        if col_name != "checkbox":
            return
        iid = self.status_tree.identify_row(event.y)
        if not iid or iid not in self.status_key_by_iid:
            return
        key = self.status_key_by_iid[iid]
        if key in self.status_checked_keys:
            self.status_checked_keys.remove(key)
        else:
            self.status_checked_keys.add(key)
        self.refresh_status_table()

    @staticmethod
    def sort_any_tree(tree, col: str, reverse: bool) -> None:
        data = [(tree.set(k, col), k) for k in tree.get_children("")]
        data.sort(reverse=reverse)
        for index, (_val, k) in enumerate(data):
            tree.move(k, "", index)
        tree.heading(col, command=lambda: GFHApp.sort_any_tree(tree, col, not reverse))

    def matching_audit_rows(self) -> List[VarianceRow]:
        rows = self.db.get_rows_by_keys(self.loaded_keys) if self.loaded_keys else []
        if not self.include_cleared.get():
            rows = [row for row in rows if not row.cleared]
        district = self.audit_district_filter.get().strip()
        store = self.audit_store_filter.get().strip()
        if district and district != "All Districts":
            rows = [row for row in rows if normalize_district(row.district) == normalize_district(district)]
        if store and store != "All Stores":
            rows = [row for row in rows if row.store == store]

        any_q = safe_text(self.audit_search_any_var.get()).lower()
        district_q = safe_text(self.audit_search_district_var.get()).lower()
        store_q = safe_text(self.audit_search_store_var.get()).lower()
        product_q = safe_text(self.audit_search_product_var.get()).lower()
        imei_q = safe_text(self.audit_search_imei_var.get()).lower()
        rep_q = safe_text(self.audit_search_rep_var.get()).lower()

        if district_q:
            rows = [row for row in rows if district_q in safe_text(row.district).lower()]
        if store_q:
            rows = [row for row in rows if store_q in safe_text(row.store).lower()]
        if product_q:
            rows = [row for row in rows if product_q in safe_text(row.product).lower()]
        if imei_q:
            rows = [row for row in rows if imei_q in safe_text(row.imei).lower()]
        if rep_q:
            rows = [row for row in rows if rep_q in safe_text(row.rep_name).lower()]
        if any_q:
            rows = [
                row for row in rows
                if any_q in " ".join([
                    safe_text(row.district),
                    safe_text(row.store),
                    safe_text(row.product),
                    safe_text(row.imei),
                    safe_text(row.status),
                    safe_text(row.rep_name),
                ]).lower()
            ]

        rows = self.filter_excluded_variance_rows(rows)
        return rows

    def auto_check_audit_rows(self) -> None:
        rows = self.matching_audit_rows()
        self.audit_checked_keys = {r.key for r in rows}
        self.refresh_table()
        self.set_status(f"Prepared {len(rows)} variance row(s) for sending. Uncheck any row you want to skip.")

    def clear_audit_checkmarks(self) -> None:
        self.audit_checked_keys.clear()
        self.refresh_table()
        self.set_status("Variance checkmarks cleared.")

    def clear_audit_search(self) -> None:
        self.audit_search_any_var.set("")
        self.audit_search_district_var.set("")
        self.audit_search_store_var.set("")
        self.audit_search_product_var.set("")
        self.audit_search_imei_var.set("")
        self.audit_search_rep_var.set("")
        self.refresh_table()

    def copy_text_to_clipboard(self, text_value: str, status_message: str = "Copied to clipboard.") -> None:
        text_value = safe_text(text_value)
        if not text_value:
            return
        try:
            self.clipboard_clear()
            self.clipboard_append(text_value)
            self.update()
            self.set_status(status_message)
        except Exception as exc:
            messagebox.showerror("Copy failed", str(exc))

    def copy_selected_imei(self, event=None) -> None:
        selected = self.audit_tree.selection() if hasattr(self, "audit_tree") else []
        if not selected:
            messagebox.showinfo("No selection", "Select a variance row first.")
            return "break"
        imeis: List[str] = []
        for iid in selected:
            values = self.audit_tree.item(iid, "values")
            if len(values) >= 4:
                imei = safe_text(values[3])
                if imei:
                    imeis.append(imei)
        if not imeis:
            messagebox.showinfo("No IMEI", "No IMEI found in selected row.")
            return "break"
        self.copy_text_to_clipboard("\n".join(imeis), f"Copied {len(imeis)} IMEI(s) to clipboard.")
        return "break"

    @staticmethod
    def audit_row_tag(row: VarianceRow) -> str:
        if row.cleared:
            return "variance_cleared"
        if row.sent_count:
            return "variance_sent"
        return "variance_pending"

    def refresh_table(self) -> None:
        for item in self.audit_tree.get_children():
            self.audit_tree.delete(item)
        self.key_by_iid.clear()

        if not self.data_loaded:
            self.summary_text.set("No data loaded")
            self.set_status("UI cleared. Select both file locations, then click Load Variances.")
            return

        rows = self.matching_audit_rows()
        rows.sort(key=lambda r: (r.district, r.store, r.rep_name, r.status, r.product, r.imei))

        for row in rows:
            clearance = "Cleared" if row.cleared else ("Sent" if row.sent_count else "Not cleared")
            checkbox = "☑ Send" if row.key in self.audit_checked_keys else "☐ Skip"
            iid = self.audit_tree.insert("", "end", values=(row.district, row.store, row.product, row.imei, row.status, row.rep_name, clearance, checkbox), tags=(self.audit_row_tag(row),))
            self.key_by_iid[iid] = row.key
        cleared_count = sum(1 for r in rows if r.cleared)
        pending_count = sum(1 for r in rows if not r.cleared)
        checked = sum(1 for r in rows if r.key in self.audit_checked_keys)
        self.set_status(f"Variance Audit rows shown: {len(rows)}. Cleared: {cleared_count}. Not cleared: {pending_count}. Checked for send: {checked}.")

    def on_audit_double_click(self, event) -> None:
        iid = self.audit_tree.identify_row(event.y)
        if not iid or iid not in self.key_by_iid:
            return
        self.audit_tree.selection_set(iid)
        self.send_selected()


    def on_audit_tree_click(self, event) -> None:
        region = self.audit_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col_id = self.audit_tree.identify_column(event.x)
        columns = self.audit_tree["columns"]
        try:
            col_name = columns[int(col_id.replace("#", "")) - 1]
        except Exception:
            return
        iid = self.audit_tree.identify_row(event.y)
        if not iid or iid not in self.key_by_iid:
            return
        if col_name == "imei":
            values = self.audit_tree.item(iid, "values")
            if len(values) >= 4:
                self.copy_text_to_clipboard(values[3], f"Copied IMEI: {values[3]}")
            return
        if col_name != "checkbox":
            return
        key = self.key_by_iid[iid]
        if key in self.audit_checked_keys:
            self.audit_checked_keys.remove(key)
        else:
            self.audit_checked_keys.add(key)
        self.refresh_table()


    def selected_rows(self) -> List[VarianceRow]:
        keys = [self.key_by_iid[iid] for iid in self.audit_tree.selection() if iid in self.key_by_iid]
        return self.db.get_rows_by_keys(keys)

    def current_mode_key(self) -> str:
        label = self.audit_send_mode.get().strip().lower()
        if label == "store":
            return "store"
        if label in {"sales rep", "rep", "salesperson"}:
            return "rep"
        return "district"

    def current_status_mode_key(self) -> str:
        label = self.status_send_mode.get().strip().lower()
        if label == "store":
            return "store"
        if label in {"sales rep", "rep", "salesperson"}:
            return "rep"
        return "district"

    @staticmethod
    def grouped_batches(rows: List[VarianceRow], mode: str) -> List[Tuple[str, str, List[VarianceRow]]]:
        grouped: Dict[Tuple[str, str], List[VarianceRow]] = {}
        for row in rows:
            district = normalize_district(row.district or "Unknown")
            if mode == "store":
                key = (district, row.store or "Unknown Store")
            elif mode == "rep":
                key = (district, row.rep_name or "Unknown Rep")
            else:
                key = (district, district)
            grouped.setdefault(key, []).append(row)
        batches: List[Tuple[str, str, List[VarianceRow]]] = []
        for (district, value), group_rows in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1])):
            title = district if mode == "district" else f"{value} | {district}"
            batches.append((title, district, group_rows))
        return batches

    @staticmethod
    def grouped_status_batches(rows: List[InventoryStatusRow], mode: str) -> List[Tuple[str, str, List[InventoryStatusRow]]]:
        grouped: Dict[Tuple[str, str], List[InventoryStatusRow]] = {}
        for row in rows:
            district = normalize_district(row.district or "Unknown")
            if mode == "store":
                key = (district, row.store or "Unknown Store")
            elif mode == "rep":
                key = (district, row.rep_name or "Unknown Rep")
            else:
                key = (district, district)
            grouped.setdefault(key, []).append(row)
        batches: List[Tuple[str, str, List[InventoryStatusRow]]] = []
        for (district, value), group_rows in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1])):
            title = district if mode == "district" else f"{value} | {district}"
            batches.append((title, district, group_rows))
        return batches

    @staticmethod
    def chunks(rows: List, size: int) -> Iterable[List]:
        for i in range(0, len(rows), size):
            yield rows[i:i + size]

    def send_checked_variances(self) -> None:
        if not self.data_loaded:
            messagebox.showinfo("Load files first", "Select both file locations and click Load Variances first.")
            return
        rows = [row for row in self.db.get_rows_by_keys(self.loaded_keys) if row.key in self.audit_checked_keys]
        rows = [row for row in rows if self.include_cleared.get() or not row.cleared]
        rows = self.filter_excluded_variance_rows(rows)
        if not rows:
            messagebox.showinfo("Nothing checked", "Select a district or store so variance rows get checked, then send them.")
            return
        mode = self.current_mode_key()
        batches = self.grouped_batches(rows, mode)
        if not messagebox.askyesno("Send checked variances", f"Send {len(rows)} checked variance row(s) in {len(batches)} image batch(es)?"):
            return
        self._send_rows_thread(rows, mode=mode, manual=True)

    def send_selected(self) -> None:
        rows = self.filter_excluded_variance_rows(self.selected_rows())
        if not rows:
            messagebox.showinfo("No selection", "Select one or more non-excluded variance rows first.")
            return
        mode = self.current_mode_key()
        self._send_rows_thread(rows, mode=mode, manual=True)

    def send_pending(self) -> None:
        if not self.data_loaded:
            messagebox.showinfo("Load files first", "Select both file locations and click Load Variances first.")
            return
        mode = self.current_mode_key()
        rows = [row for row in self.db.get_rows_by_keys(self.loaded_keys) if not row.cleared]
        district = self.audit_district_filter.get().strip()
        store = self.audit_store_filter.get().strip()
        if district and district != "All Districts":
            rows = [row for row in rows if normalize_district(row.district) == normalize_district(district)]
        if store and store != "All Stores":
            rows = [row for row in rows if row.store == store]
        rows = self.filter_excluded_variance_rows(rows)
        if self.send_only_unsent.get():
            rows = [row for row in rows if not row.last_sent_at]
        if not rows:
            messagebox.showinfo("Nothing to send", "No pending unsent variances found in the current filter.")
            return
        label = SEND_MODE_LABELS.get(mode, mode.title())
        batches = self.grouped_batches(rows, mode)
        if not messagebox.askyesno("Send pending", f"Send {len(rows)} pending variance rows in {len(batches)} image batch(es) by {label}?\n\nImages will route to district WhatsApp groups."):
            return
        self._send_rows_thread(rows, mode=mode, manual=False)

    def send_checked_status(self) -> None:
        if not self.data_loaded:
            messagebox.showinfo("Load files first", "Select both file locations and click Load Variances first.")
            return
        rows = [self.status_row_by_key[k] for k in self.status_checked_keys if k in self.status_row_by_key]
        if not rows:
            messagebox.showinfo("Nothing checked", "Select a district or store so Inventory Audit Status rows get checked, then send them.")
            return
        mode = self.current_status_mode_key()
        batches = self.grouped_status_batches(rows, mode)
        if not messagebox.askyesno("Send Inventory Audit Status", f"Send {len(rows)} inventory status row(s) in {len(batches)} image batch(es)?"):
            return
        thread = threading.Thread(target=self._send_status_rows, args=(rows, mode), daemon=True)
        thread.start()

    def inventory_status_message(self) -> str:
        today = dt.datetime.now()
        return f"Inventory Audit Status as of {today.month}/{today.day}/{today.year}."

    def variance_request_message(self, rows: List[VarianceRow]) -> str:
        seen_phones: set[str] = set()
        phones: List[str] = []
        missing_names: List[str] = []

        for row in rows:
            rep_name = safe_text(row.rep_name)
            if not rep_name:
                continue

            phone = normalize_phone(self.db.find_sales_rep_phone(rep_name))
            if not phone:
                missing_names.append(rep_name)
                continue

            if phone in seen_phones:
                continue

            seen_phones.add(phone)
            phones.append(whatsapp_mention(phone))

        if missing_names:
            missing_unique = sorted(set(missing_names))
            self.set_status(
                "No employee phone match found for: " + ", ".join(missing_unique[:5]) +
                ("..." if len(missing_unique) > 5 else "")
            )

        if not phones:
            return ""

        return " ".join(phones) + " please share the images of the variances."


    def pending_inventory_count_message(self, rows: List[InventoryStatusRow]) -> str:
        messages: List[str] = []
        seen: set[str] = set()

        for row in rows:
            if safe_text(row.status).lower() == "completed":
                continue

            rep_name = safe_text(row.rep_name)
            phone = normalize_phone(self.db.find_sales_rep_phone(rep_name)) if rep_name else ""
            if phone:
                target = whatsapp_mention(phone)
            else:
                target = safe_text(row.store)

            if not target or target in seen:
                continue

            seen.add(target)
            messages.append(f"{target}, please complete an inventory count ASAP")

        return "\n".join(messages)

    def create_full_inventory_audit_log(self, final_results: List[Dict[str, str]]) -> Path:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = EXPORT_DIR / f"GFH_Full_Inventory_Audit_Log_{stamp}.xlsx"
        variance_rows = self.db.get_rows_by_keys(self.loaded_keys) if self.loaded_keys else []
        final_results = final_results or []

        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Section", "District", "Store", "Product", "IMEI", "Status", "Rep Name",
            "Clearance Status", "Sent Count", "Last Sent At", "Final Caption",
            "Pending Inventory Stores", "Final Image Path", "WhatsApp Group", "Generated At"
        ]
        ws.append(headers)
        
        imei_col_idx = headers.index("IMEI") + 1
        
        current_row = 2

        for row in sorted(self.status_rows, key=lambda r: (r.district, r.store, r.rep_name)):
            data = [
                "Inventory Status", row.district, row.store, "", "", row.status, row.rep_name,
                "", "", "", "", "", "", "", now_text()
            ]
            for c_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                if c_idx == imei_col_idx:
                    cell.data_type = 's'
                    cell.number_format = '@'
            current_row += 1

        for row in sorted(variance_rows, key=lambda r: (r.district, r.store, r.rep_name, r.product, r.imei)):
            data = [
                "Variance Audit", row.district, row.store, row.product, row.imei, row.status, row.rep_name,
                "Cleared" if row.cleared else "Not Cleared", row.sent_count, row.last_sent_at,
                "", "", "", "", now_text()
            ]
            for c_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                if c_idx == imei_col_idx:
                    cell.data_type = 's'
                    cell.number_format = '@'
            current_row += 1

        for item in final_results:
            data = [
                "Final District Send", item.get("district", ""), item.get("stores", ""), "", "",
                item.get("status", ""), "", "", "", "", item.get("caption", ""),
                item.get("pending_inventory_stores", ""), item.get("image_path", ""),
                item.get("group_name", ""), item.get("sent_at", "")
            ]
            for c_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                if c_idx == imei_col_idx:
                    cell.data_type = 's'
                    cell.number_format = '@'
            current_row += 1

        wb.save(output_path)
        return output_path

    def _send_rows_thread(self, rows: List[VarianceRow], mode: str, manual: bool) -> None:
        thread = threading.Thread(target=self._send_rows, args=(rows, mode, manual), daemon=True)
        thread.start()

    def _send_rows(self, rows: List[VarianceRow], mode: str, manual: bool) -> None:
        try:
            renderer = ImageRenderer()
            sender = WhatsAppSender(status_callback=self.set_status)
            batches = self.grouped_batches(rows, mode)
            send_mode_for_file = mode
            for batch_title, district, batch_rows in batches:
                group_name = group_name_for_district(district, self.db)
                for chunk_no, chunk_rows in enumerate(self.chunks(batch_rows, AUTO_SEND_CHUNK_SIZE), start=1):
                    title = batch_title if len(batch_rows) <= AUTO_SEND_CHUNK_SIZE else f"{batch_title} | Part {chunk_no}"
                    image_path = renderer.render_rows(title, chunk_rows, mode=send_mode_for_file)
                    try:
                        self.set_status(f"Sending {len(chunk_rows)} variance row(s) to {group_name}. {title}.")
                        sender.send_image(group_name, image_path, text_message=self.variance_request_message(chunk_rows))
                        self.db.mark_sent(chunk_rows, group_name, title, str(image_path), mode=send_mode_for_file)
                    except Exception as exc:
                        self.db.mark_sent(chunk_rows, group_name, title, str(image_path), mode=send_mode_for_file, error=str(exc))
                        messagebox.showerror("WhatsApp send failed", f"Image was created but WhatsApp send failed.\n\nGroup: {group_name}\nBatch: {title}\nImage: {image_path}\n\nError: {exc}")
                        self.set_status(f"Send failed. Image saved: {image_path}")
                        return
            self.refresh_table()
            self.set_status("Variance image sending completed.")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Send failed", str(exc))
            self.set_status("Send failed.")

    def _render_status_rows(self, batch_title: str, rows: List[InventoryStatusRow], mode: str = "status") -> Path:
        if Image is None:
            raise RuntimeError("Pillow is required. Install with: py -m pip install pillow")

        renderer = ImageRenderer()
        width = 1320
        margin = 28
        title_font = renderer._font(28, True)
        header_font = renderer._font(16, True)
        cell_font = renderer._font(15, False)
        cell_bold = renderer._font(15, True)
        sub_font = renderer._font(15, False)
        row_h = 34
        black = (18, 20, 43)
        border = (0, 0, 0)
        light_green = (210, 240, 214)
        pending_fill = (238, 230, 177)
        white = (255, 255, 255)
        district_colors = {
            "Arizona": (196, 138, 230),
            "Colorado East": (96, 194, 236),
            "Colorado West": (78, 186, 228),
            "Houston": (244, 128, 128),
            "Louisiana": (247, 199, 34),
            "Tennessee": (74, 145, 62),
        }

        logo_img = None
        logo_w = 0
        logo_h_used = 0
        if STATUS_LOGO_PATH.exists():
            try:
                logo_img = Image.open(STATUS_LOGO_PATH).convert("RGBA")
                scale = min(560 / logo_img.width, 135 / logo_img.height)
                size = (max(1, int(logo_img.width * scale)), max(1, int(logo_img.height * scale)))
                logo_img = logo_img.resize(size)
                logo_w, logo_h_used = size
            except Exception:
                logo_img = None
                logo_w = 0
                logo_h_used = 0

        header_area_h = max(logo_h_used, 80) + 18
        table_top = margin + header_area_h + 18
        height = table_top + 26 + row_h * len(rows) + 24

        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)

        title_text = "GFH Inventory Audit"
        today = dt.datetime.now()
        date_text = f"{today.month}/{today.day}/{today.year}"

        # Measure each string so we can center it horizontally
        title_bbox = draw.textbbox((0, 0), title_text, font=title_font)
        title_w = title_bbox[2] - title_bbox[0]
        date_bbox = draw.textbbox((0, 0), date_text, font=cell_bold)
        date_w = date_bbox[2] - date_bbox[0]

        title_x = (width - title_w) // 2
        title_y = margin + 6
        draw.text((title_x, title_y), title_text, fill=black, font=title_font)

        date_x = (width - date_w) // 2
        draw.text((date_x, title_y + 40), date_text, fill=black, font=cell_bold)

        if safe_text(batch_title):
            bt_text = safe_text(batch_title)
            bt_bbox = draw.textbbox((0, 0), bt_text, font=sub_font)
            bt_w = bt_bbox[2] - bt_bbox[0]
            bt_x = (width - bt_w) // 2
            draw.text((bt_x, title_y + 40 + 26), bt_text, fill=black, font=sub_font)

        if logo_img is not None:
            logo_x = width - margin - logo_w
            logo_y = margin
            img.paste(logo_img, (logo_x, logo_y), logo_img)

        y = table_top
        x = margin
        col_widths = [160, 280, 145, 679]
        headers = ["District", "Store", "Status", "Salesperson"]
        draw.rectangle((x, y, width - margin, y + 26), outline=border, fill=white)
        cx = x
        for idx, h in enumerate(headers):
            draw.line((cx, y, cx, y + 26), fill=border, width=1)
            draw.text((cx + 4, y + 4), h, fill=black, font=header_font)
            cx += col_widths[idx]
        draw.line((width - margin, y, width - margin, y + 26), fill=border, width=1)
        draw.line((x, y + 26, width - margin, y + 26), fill=border, width=1)
        y += 26

        rows_sorted = sorted(rows, key=lambda r: (normalize_district(r.district), r.store))
        for row in rows_sorted:
            district = normalize_district(row.district)
            row_fill = district_colors.get(district, (245, 245, 245))
            draw.rectangle((x, y, width - margin, y + row_h), fill=row_fill, outline=border)
            vals = [district, row.store, row.status, row.rep_name or "-"]
            cx = x
            for cidx, val in enumerate(vals):
                if cidx == 2:
                    status_fill = light_green if safe_text(row.status).lower() == "completed" else pending_fill
                    draw.rectangle((cx, y, cx + col_widths[cidx], y + row_h), fill=status_fill, outline=border)
                draw.line((cx, y, cx, y + row_h), fill=border, width=1)
                text_val = safe_text(val) or "-"
                if len(text_val) > 42 and cidx == 3:
                    text_val = text_val[:39] + "..."
                if len(text_val) > 22 and cidx == 1:
                    text_val = text_val[:19] + "..."
                font = cell_bold if cidx in (0, 1, 3) else cell_font
                fill = (210, 0, 0) if (cidx == 2 and safe_text(row.status).lower() != "completed") else black
                draw.text((cx + 4, y + 6), text_val, fill=fill, font=font)
                cx += col_widths[cidx]
            draw.line((width - margin, y, width - margin, y + row_h), fill=border, width=1)
            y += row_h

        safe_title = re.sub(r"[^A-Za-z0-9_-]+", "_", batch_title)[:60] or "Inventory_Audit_Status"
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = IMAGE_DIR / f"GFH_Inventory_Audit_Status_{safe_title}_{stamp}.png"
        img.save(path)
        return path

    def _send_status_rows(self, rows: List[InventoryStatusRow], mode: str) -> None:
        try:
            sender = WhatsAppSender(status_callback=self.set_status)
            batches = self.grouped_status_batches(rows, mode)
            for batch_title, district, batch_rows in batches:
                group_name = group_name_for_district(district, self.db)
                for chunk_no, chunk_rows in enumerate(self.chunks(batch_rows, 28), start=1):
                    title = batch_title if len(batch_rows) <= 28 else f"{batch_title} | Part {chunk_no}"
                    image_path = self._render_status_rows(title, chunk_rows, mode="inventory status")
                    self.set_status(f"Sending Inventory Audit Status with {len(chunk_rows)} row(s) to {group_name}. {title}.")
                    sender.send_image(group_name, image_path, text_message=self.inventory_status_message())
                    self.db.mark_status_sent(chunk_rows)
                    pending_message = self.pending_inventory_count_message(chunk_rows)
                    if pending_message:
                        self.set_status(f"Sending incomplete inventory count reminder to {group_name}.")
                        sender.send_text(group_name, pending_message)
            self.refresh_status_table()
            self.set_status("Inventory Audit Status image sending completed.")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Send failed", str(exc))
            self.set_status("Inventory Audit Status send failed.")

    @staticmethod
    def format_store_list_caption(stores: List[str]) -> str:
        clean = [safe_text(s) for s in stores if safe_text(s)]
        if not clean:
            return ""
        if len(clean) == 1:
            return clean[0]
        if len(clean) == 2:
            return f"{clean[0]} and {clean[1]}"
        return ", ".join(clean[:-1]) + f", and {clean[-1]}"

    def final_district_caption(self, district: str, rows: List[VarianceRow], pending_status_rows: Optional[List[InventoryStatusRow]] = None) -> str:
        parts: List[str] = []
        dm_phone = self.db.find_district_manager_phone(district)
        if dm_phone:
            parts.append(whatsapp_mention(dm_phone))

        stores = sorted({safe_text(r.store) for r in rows if safe_text(r.store)})
        if stores:
            store_text = self.format_store_list_caption(stores)
            parts.append(f"An Inventory audit has been completed, variance found in {store_text}.")
        else:
            parts.append("An Inventory audit has been completed and no variance found.")

        pending_stores = sorted({safe_text(r.store) for r in (pending_status_rows or []) if safe_text(r.store)})
        if pending_stores:
            pending_text = self.format_store_list_caption(pending_stores)
            verb = "did not complete" if len(pending_stores) == 1 else "did not complete"
            parts.append(f"{pending_text} {verb} an inventory count.")

        return " ".join(parts)

    def _render_no_variance_image(self, district: str) -> Path:
        if Image is None:
            raise RuntimeError("Pillow is required. Install with: py -m pip install pillow")

        renderer = ImageRenderer()
        width = 1360
        margin = 32
        title_font = renderer._font(34, True)
        sub_font = renderer._font(20, False)
        body_font = renderer._font(28, True)
        note_font = renderer._font(18, False)

        dark = (18, 20, 43)
        gray = (95, 95, 102)
        red = (233, 27, 47)
        border = (215, 218, 223)

        logo_img = None
        logo_w = 0
        logo_h_used = 0
        if STATUS_LOGO_PATH.exists():
            try:
                logo_img = Image.open(STATUS_LOGO_PATH).convert("RGBA")
                scale = min(620 / logo_img.width, 150 / logo_img.height)
                size = (max(1, int(logo_img.width * scale)), max(1, int(logo_img.height * scale)))
                logo_img = logo_img.resize(size)
                logo_w, logo_h_used = size
            except Exception:
                logo_img = None
                logo_w = 0
                logo_h_used = 0

        header_area_h = max(logo_h_used, 74) + 18
        box_top = margin + header_area_h + 16
        height = box_top + 190 + 40
        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)

        draw.text((margin, margin + 4), "GFH Inventory Variance", fill=dark, font=title_font)
        draw.text((margin, margin + 48), f"{district}   Generated: {now_text()}", fill=gray, font=sub_font)

        if logo_img is not None:
            logo_x = width - margin - logo_w
            logo_y = margin
            img.paste(logo_img, (logo_x, logo_y), logo_img)

        draw.rectangle((margin, box_top, width - margin, box_top + 160), outline=border, fill=(255, 255, 255))
        draw.rectangle((margin, box_top, width - margin, box_top + 46), fill=red)
        draw.text((margin + 14, box_top + 11), "Final Audit Result", fill="white", font=sub_font)
        draw.text((margin + 18, box_top + 78), "No variance found in this district.", fill=dark, font=body_font)
        draw.text((margin + 18, box_top + 118), "All variances were cleared or none were found.", fill=gray, font=note_font)

        safe_title = re.sub(r"[^A-Za-z0-9_-]+", "_", district)[:60] or "District"
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = IMAGE_DIR / f"GFH_Final_Audit_{safe_title}_{stamp}.png"
        img.save(path)
        return path

    def send_starting_message(self) -> None:
        """Send 'Please complete an Inventory count in 15 minutes.' to district WhatsApp groups."""
        if not self.data_loaded:
            messagebox.showinfo("Load files first", "Select both file locations and click Load Variances first.")
            return
        all_districts = sorted({normalize_district(row.district) for row in self.status_rows if safe_text(row.district)})
        if not all_districts:
            messagebox.showinfo("No districts", "No district data is available.")
            return
        selected = self.final_district_var.get().strip()
        if selected and selected != "All Districts":
            districts = [normalize_district(selected)]
        else:
            districts = all_districts
        if not messagebox.askyesno("Send starting message", f"Send starting message to {len(districts)} district group(s)?\n\nMessage: Please complete an Inventory count in 15 minutes."):
            return
        thread = threading.Thread(target=self._send_starting_message_thread, args=(districts,), daemon=True)
        thread.start()

    def _send_starting_message_thread(self, districts: List[str]) -> None:
        try:
            sender = WhatsAppSender(status_callback=self.set_status)
            message = "Please complete an Inventory count in 15 minutes."
            for district in districts:
                group_name = group_name_for_district(district, self.db)
                self.set_status(f"Sending starting message to {group_name}...")
                sender.send_text(group_name, message)
            self.set_status("Starting message sent to all selected districts.")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Send failed", str(exc))
            self.set_status(f"Starting message send failed: {exc}")

    def send_reminder(self, reminder_number: int) -> None:
        """Send a single reminder message for uncleared variances to district WhatsApp groups."""
        reminder_messages = {
            1: "Please clear the pending variances.",
            2: "This is the second reminder. Please clear the pending variances immediately.",
            3: "Final reminder. Uncleared variances will be reported.",
        }
        if reminder_number not in reminder_messages:
            return
        if not self.data_loaded:
            messagebox.showinfo("Load files first", "Select both file locations and click Load Variances first.")
            return
        all_rows = self.db.get_rows_by_keys(self.loaded_keys)
        uncleared_rows = [row for row in all_rows if not row.cleared]
        uncleared_rows = self.filter_excluded_variance_rows(uncleared_rows)
        if not uncleared_rows:
            messagebox.showinfo("No uncleared variances", "All variances have been cleared. No reminder needed.")
            return
        all_districts = sorted({normalize_district(row.district) for row in uncleared_rows if safe_text(row.district)})
        if not all_districts:
            messagebox.showinfo("No districts", "No district data available for reminders.")
            return
        selected = self.final_district_var.get().strip()
        if selected and selected != "All Districts":
            districts = [normalize_district(selected)]
        else:
            districts = all_districts
        message = reminder_messages[reminder_number]
        if not messagebox.askyesno(
            f"Send Reminder {reminder_number}",
            f"Send reminder {reminder_number} to {len(districts)} district group(s)?\n\n"
            f"Message: {message}\n\n"
            f"Total {len(uncleared_rows)} uncleared variance row(s).",
        ):
            return
        thread = threading.Thread(target=self._send_single_reminder_thread, args=(districts, uncleared_rows, reminder_number, message), daemon=True)
        thread.start()

    def _send_single_reminder_thread(self, districts: List[str], uncleared_rows: List[VarianceRow], reminder_number: int, message: str) -> None:
        try:
            sender = WhatsAppSender(status_callback=self.set_status)
            for district in districts:
                district_rows = [row for row in uncleared_rows if normalize_district(row.district) == normalize_district(district)]
                if not district_rows:
                    continue
                group_name = group_name_for_district(district, self.db)
                rep_mentions = self.variance_request_message(district_rows)
                full_message = f"Reminder {reminder_number}/3: {message}"
                if rep_mentions:
                    full_message = f"{rep_mentions} {message}"
                self.set_status(f"Sending reminder {reminder_number}/3 to {group_name}...")
                sender.send_text(group_name, full_message)
                time.sleep(2)
            self.set_status(f"Reminder {reminder_number} sent successfully.")
            try:
                messagebox.showinfo("Reminder sent", f"Reminder {reminder_number} has been sent to the selected district group(s).")
            except Exception:
                pass
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Send failed", str(exc))
            self.set_status(f"Reminder {reminder_number} send failed: {exc}")

    def send_final_district_result(self) -> None:
        if not self.data_loaded:
            messagebox.showinfo("Load files first", "Select both file locations and click Load Variances first.")
            return
        all_districts = sorted({normalize_district(row.district) for row in self.status_rows if safe_text(row.district)})
        if not all_districts:
            all_districts = sorted({normalize_district(row.district) for row in self.db.get_rows_by_keys(self.loaded_keys) if safe_text(row.district)})
        if not all_districts:
            messagebox.showinfo("No districts", "No district data is available for final sending.")
            return
        selected = self.final_district_var.get().strip()
        if selected and selected != "All Districts":
            districts = [normalize_district(selected)]
        else:
            districts = all_districts
        if not messagebox.askyesno("Send final district result", f"Send final audit result to {len(districts)} district group(s)?"):
            return
        thread = threading.Thread(target=self._send_final_district_result_thread, args=(districts,), daemon=True)
        thread.start()

    def _send_final_district_result_thread(self, districts: List[str]) -> None:
        final_results: List[Dict[str, str]] = []
        try:
            sender = WhatsAppSender(self.set_status)
            renderer = ImageRenderer()
            all_rows = self.db.get_rows_by_keys(self.loaded_keys)
            self.set_status("Sending final district audit results...")
            for district in districts:
                district_rows = [row for row in all_rows if normalize_district(row.district) == normalize_district(district)]
                district_rows = [row for row in district_rows if not row.cleared]
                district_rows = self.filter_excluded_variance_rows(district_rows)
                group_name = group_name_for_district(district, self.db)
                pending_status_rows = [
                    row for row in self.status_rows
                    if normalize_district(row.district) == normalize_district(district)
                    and normalize_header(row.status) != "completed"
                ]
                caption = self.final_district_caption(district, district_rows, pending_status_rows)
                if district_rows:
                    image_path = renderer.render_rows(district, district_rows, mode="district")
                    status = "Variance Found"
                else:
                    image_path = self._render_no_variance_image(district)
                    status = "No Variance Found"
                self.set_status(f"Sending final district result to {group_name}.")
                sender.send_image(group_name, image_path, text_message=caption)
                final_results.append({
                    "district": district,
                    "stores": self.format_store_list_caption(sorted({safe_text(r.store) for r in district_rows if safe_text(r.store)})),
                    "pending_inventory_stores": self.format_store_list_caption(sorted({safe_text(r.store) for r in pending_status_rows if safe_text(r.store)})),
                    "status": status,
                    "caption": caption,
                    "image_path": str(image_path),
                    "group_name": group_name,
                    "sent_at": now_text(),
                })

            log_path = self.create_full_inventory_audit_log(final_results)
            self.set_status(f"Final district audit result sending completed. Full audit log created: {log_path}")
            try:
                messagebox.showinfo("Audit log created", f"Full inventory audit log created:\n{log_path}")
            except Exception:
                pass
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Send failed", str(exc))
            self.set_status(f"Final district sending failed: {exc}")

    def mark_selected(self, cleared: bool) -> None:
        rows = self.selected_rows()
        if not rows:
            messagebox.showinfo("No selection", "Select one or more variance rows first.")
            return
        def _do():
            for row in rows:
                self.db.set_cleared(row.key, cleared)
        self._db_write(_do)
        self.refresh_table()
        action = "cleared" if cleared else "not cleared"
        self.set_status(f"Marked {len(rows)} variance row(s) as {action}.")

    def clear_current_ui(self, silent: bool = False) -> None:
        self.loaded_keys.clear()
        self.data_loaded = False
        self.status_rows = []
        self.status_row_by_key = {}
        self.status_checked_keys.clear()
        self.audit_checked_keys.clear()
        self.status_key_by_iid.clear()
        self.populate_status_filters()
        self.populate_audit_filters()
        self.refresh_status_table()
        self.refresh_table()
        if not silent:
            self.set_status("UI cleared. Select both file locations, then click Load Variances.")

    def export_log(self) -> None:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_path = EXPORT_DIR / f"GFH_Variance_Clearance_Log_{stamp}.xlsx"
        path = filedialog.asksaveasfilename(title="Export Clearance Log", initialdir=str(EXPORT_DIR), initialfile=default_path.name, defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not path:
            return
        keys = self.loaded_keys if self.data_loaded else None
        self.db.export_xlsx(Path(path), keys=keys)
        self.set_status(f"Exported clearance log: {path}")
        messagebox.showinfo("Export complete", f"Log exported:\n{path}")

    # ── Cross-instance DB sync ───────────────────────────────────────────────
    def _start_db_sync_poll(self) -> None:
        """Begin the 2-second mtime-polling loop."""
        self._schedule_db_sync()

    def _schedule_db_sync(self) -> None:
        self.after(2000, self._check_db_sync)

    def _check_db_sync(self) -> None:
        """Called every 2 s.  If another instance modified the DB, refresh GUI."""
        try:
            if not self._db_sync_paused and DB_PATH.exists():
                mtime = DB_PATH.stat().st_mtime
                if mtime != self._db_mtime:
                    self._db_mtime = mtime
                    # Reload master data in case store/rep lists changed too
                    try:
                        self.master_store_records = self.db.store_master_records()
                    except Exception:
                        pass
                    # Refresh both visible tables silently
                    try:
                        self.refresh_table()
                    except Exception:
                        pass
                    try:
                        self.refresh_status_table()
                    except Exception:
                        pass
        except Exception:
            pass
        self._schedule_db_sync()

    def _db_write(self, fn, *args, **kwargs):
        """Wrap any DB write so the sync poller does not re-trigger on our own write."""
        self._db_sync_paused = True
        try:
            result = fn(*args, **kwargs)
        finally:
            # Update our own mtime baseline after writing, then re-enable polling
            try:
                self._db_mtime = DB_PATH.stat().st_mtime
            except Exception:
                pass
            self._db_sync_paused = False
        return result

    def open_app_folder(self) -> None:
        APP_DIR.mkdir(parents=True, exist_ok=True)
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(APP_DIR))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(APP_DIR)])
            else:
                subprocess.Popen(["xdg-open", str(APP_DIR)])
        except Exception as exc:
            messagebox.showerror("Open folder failed", str(exc))

def _enable_dpi_awareness() -> None:
    """Make Windows report physical pixels so winfo_screen* is accurate on
    high-DPI displays (1080p, 1440p, 2K, 4K, DPI-scaled laptops)."""
    if sys.platform != "win32":
        return
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)  # system DPI aware
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


def main() -> None:
    _enable_dpi_awareness()
    app = GFHApp()
    app.mainloop()


if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    try:
        main()
    except Exception as exc:
        traceback.print_exc()
        show_startup_error(exc)
        if sys.stdin and sys.stdin.isatty():
            input("Press Enter to close...")